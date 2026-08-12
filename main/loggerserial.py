"""
logger_serial_ke_excel.py

TUJUAN:
Membaca SEMUA baris yang dikirim ESP32-S3 lewat Serial (USB):
  1. Ditampilkan LANGSUNG di terminal, mirip Serial Monitor Arduino IDE,
     dengan warna berbeda per jenis baris (JSON, WARNING, FFT-DIAG, dst)
     supaya gampang dipantau mata sambil alat menyala.
  2. Baris yang berupa JSON data sensor (diawali karakter '{') SEKALIGUS
     ditulis sebagai baris baru ke file Excel (.xlsx) dan CSV cadangan,
     real-time, tanpa perlu copy-paste manual.

CARA PAKAI:
1. Install dependency (sekali saja):
       pip install pyserial openpyxl colorama
2. Tutup Arduino IDE Serial Monitor dulu (satu port cuma bisa dipakai
   satu program dalam satu waktu -- kalau Serial Monitor Arduino IDE
   masih terbuka, script ini akan gagal connect).
3. Cek nama port ESP32 kamu di Device Manager (Windows), biasanya
   berbentuk "COM3", "COM5", dst. Update variabel SERIAL_PORT di bawah.
4. Jalankan: python logger_serial_ke_excel.py
5. Biarkan berjalan selama alat menyala. Semua baris tampil di terminal;
   baris JSON otomatis ditambahkan sebagai baris baru di file Excel.
6. Tekan CTRL+C untuk berhenti. File Excel aman tersimpan.

CATATAN PENTING:
- Selama script ini berjalan, JANGAN buka file Excel-nya di Excel
  secara bersamaan di Windows -- Excel mengunci file saat dibuka,
  script akan gagal menulis. Tutup dulu file Excel-nya, baru jalankan
  script, atau buka file itu SETELAH kamu berhenti logging.
- Kalau ingin melihat data yang sedang masuk tanpa menutup logger,
  buka file CSV pendampingnya (dibuat otomatis, nama sama, akhiran
  .csv) -- CSV tidak dikunci Excel selama tidak dibuka juga.
"""

import serial
import json
import csv
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from colorama import init as colorama_init, Fore, Style

colorama_init(autoreset=True)  # supaya warna jalan juga di Command Prompt Windows

# ======================= KONFIGURASI =======================
SERIAL_PORT = "COM5"      # GANTI sesuai port ESP32 kamu (lihat Device Manager)
BAUD_RATE = 115200        # Harus sama dengan Serial.begin() di main.ino

# Isi manual SEBELUM tiap sesi jalan -- ini yang bikin nama file gampang
# dibaca ulang bulan depan tanpa perlu inget urutan tes.
KONDISI_TES = "motorSteady"   # contoh: motorOFF, motorSteady, faultUnbalance, dst.

_timestamp = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT_XLSX = f"vibris_{_timestamp}_{KONDISI_TES}.xlsx"
OUTPUT_CSV  = f"vibris_{_timestamp}_{KONDISI_TES}.csv"
# =============================================================

# Urutan kolom mengikuti field JSON yang dikirim RaspberryPiDataTransmitter.cpp
KOLOM = [
    "waktu_lokal", "rms_v", "rms_x", "rms_y", "rms_z", "rms_a",
    "cur", "cur_raw_adc", "temp", "temp_raw", "rpm", "snr", "severity",
    "status", "e_unbalance", "e_misalign", "e_bpfo", "e_bpfi",
    "diagnosis", "diag_conf",
    "e_audio_low", "e_audio_mid", "e_audio_high",
    "audio_diagnosis", "audio_diag_conf",
    "ml_label", "ml_conf",
    "roughness", "brightness",
    "ground_truth"
]


def warna_untuk(baris_teks):
    """Tentukan warna tampilan terminal berdasarkan jenis baris -- meniru
    kode warna yang sama dipakai di versi Word log kemarin, supaya kamu
    gak perlu belajar skema baru: WARNING=merah, FFT-DIAG=biru,
    FFT=hijau, blok TELEMETRI=magenta, JSON=putih redup, lainnya=default."""
    if baris_teks.startswith("[WARNING]"):
        return Fore.RED
    if "[FFT-DIAG]" in baris_teks:
        return Fore.CYAN
    if baris_teks.startswith("[FFT]"):
        return Fore.GREEN
    if baris_teks.startswith("="):
        return Fore.MAGENTA
    if baris_teks.startswith("{"):
        return Fore.WHITE + Style.DIM
    if baris_teks.startswith("[SYSTEM]") or baris_teks.startswith("[Calibrator]"):
        return Fore.YELLOW
    if "[AUDIO]" in baris_teks:
        return Fore.CYAN + Style.DIM
    return Fore.RESET


def siapkan_file_excel(path):
    """Buat file Excel baru dengan header kalau belum ada, atau lanjutkan
    dari file yang sudah ada supaya tidak menimpa data lama."""
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Telemetri"
    ws.append(KOLOM)
    wb.save(path)
    return wb, ws


def siapkan_file_csv(path):
    file_baru = not os.path.exists(path)
    f = open(path, "a", newline="", encoding="utf-8")
    writer = csv.writer(f)
    if file_baru:
        writer.writerow(KOLOM)
    return f, writer


def main():
    print(f"[LOGGER] Menyambungkan ke {SERIAL_PORT} @ {BAUD_RATE} baud...")
    try:
        ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=2)
    except serial.SerialException as e:
        print(f"[LOGGER] GAGAL membuka {SERIAL_PORT}: {e}")
        print("[LOGGER] Cek: nama port benar (lihat Device Manager)? "
              "Arduino Serial Monitor sudah ditutup? Kabel USB tersambung?")
        return  # keluar dengan pesan jelas, bukan traceback mentah

    print("[LOGGER] Tersambung. Menunggu data dari ESP32... "
          "(semua baris tampil di bawah, mirip Serial Monitor)\n")

    wb, ws = siapkan_file_excel(OUTPUT_XLSX)
    csv_file, csv_writer = siapkan_file_csv(OUTPUT_CSV)

    baris_masuk = 0
    try:
        while True:
            raw_line = ser.readline().decode("utf-8", errors="ignore").strip()

            if not raw_line:
                continue  # timeout tanpa data, jangan cetak baris kosong

            # TAMPILKAN SEMUA baris ke terminal -- ini bagian "seperti
            # Serial Monitor" yang diminta, jalan untuk SEMUA jenis baris,
            # bukan cuma yang JSON.
            warna = warna_untuk(raw_line)
            print(f"{warna}{raw_line}{Style.RESET_ALL}")

            # Baris yang bukan JSON (TELEMETRI, [FFT], [WARNING], dst) sudah
            # ditampilkan di atas -- tapi tidak masuk Excel, cuma data sensor
            # terstruktur yang disimpan.
            if not raw_line.startswith("{"):
                continue

            try:
                data = json.loads(raw_line)
            except json.JSONDecodeError:
                # Baris JSON kepotong/rusak di tengah transmisi -- lewati
                # saja, jangan sampai bikin script crash.
                print(f"{Fore.RED}[LOGGER] Baris JSON rusak, dilewati.{Style.RESET_ALL}")
                continue

            waktu = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            baris = [
                waktu,
                data.get("rms_v"), data.get("rms_x"), data.get("rms_y"), data.get("rms_z"),
                data.get("rms_a"), data.get("cur"), data.get("cur_raw_adc"),
                data.get("temp"), data.get("temp_raw"), data.get("rpm"), data.get("snr"),
                data.get("severity"), data.get("status"),
                data.get("e_unbalance"), data.get("e_misalign"),
                data.get("e_bpfo"), data.get("e_bpfi"),
                data.get("diagnosis"), data.get("diag_conf"),
                data.get("e_audio_low"), data.get("e_audio_mid"), data.get("e_audio_high"),
                data.get("audio_diagnosis"), data.get("audio_diag_conf"),
                data.get("ml_label"), data.get("ml_conf"),
                data.get("roughness"), data.get("brightness"),
                data.get("ground_truth"),
            ]

            # Tulis ke Excel
            ws.append(baris)
            wb.save(OUTPUT_XLSX)   # simpan tiap baris masuk -- aman kalau listrik/alat mati mendadak

            # Tulis ke CSV juga (cadangan, bisa dipantau live tanpa lock Excel)
            csv_writer.writerow(baris)
            csv_file.flush()

            baris_masuk += 1
            print(f"{Fore.YELLOW}[LOGGER] Baris #{baris_masuk} tersimpan ke Excel | "
                  f"status={data.get('status')} rpm={data.get('rpm')}{Style.RESET_ALL}")

    except KeyboardInterrupt:
        print("\n[LOGGER] Dihentikan oleh user. Menyimpan file terakhir...")
    except serial.SerialException as e:
        print(f"[LOGGER] ERROR koneksi serial: {e}")
        print("[LOGGER] Cek: kabel USB tercabut? Device Manager masih mendeteksi port?")
    finally:
        wb.save(OUTPUT_XLSX)
        csv_file.close()
        ser.close()
        print(f"[LOGGER] Selesai. Total {baris_masuk} baris tersimpan di {OUTPUT_XLSX} dan {OUTPUT_CSV}")


if __name__ == "__main__":
    main()