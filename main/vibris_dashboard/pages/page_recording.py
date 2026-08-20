from PyQt5.QtWidgets import (
    QWidget, QLabel, QPushButton, QHBoxLayout, QVBoxLayout,
    QGridLayout, QStackedWidget, QFrame, QListWidget, QComboBox, QMessageBox,
    QSlider, QSizePolicy, QProgressBar, QLineEdit
)
from PyQt5.QtCore import QTimer, Qt, QPoint, QEvent
from PyQt5.QtGui import QFont, QPainter, QLinearGradient, QColor, QPolygon, QPen, QBrush
import pyqtgraph as pg
import os
import csv
import time
from datetime import datetime
from config import COL_PANEL_DARK, COL_BAD, LOG_DIR
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None


class RecordingPageMixin:
    """Method-method untuk halaman ini. Di-mixin ke class Dashboard
    (lihat dashboard_core.py) supaya semua method di sini bisa akses
    self.current_v, self.selected_slot_idx, dst -- gak ada state baru di sini,
    murni pemisahan biar dashboard_core.py gak sepanjang 2000 baris."""

    def _page_recording(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(4)

        rec_ctrl = QHBoxLayout()
        self.btn_toggle_rec = QPushButton("MULAI RECORDING")
        self.btn_toggle_rec.setStyleSheet("background-color: #cfcfcf; color: #000000; font-weight: bold; font-size: 8px; height: 22px; border-radius: 4px;")
        self.btn_toggle_rec.clicked.connect(self._toggle_recording)
        
        self.lbl_rec_status = QLabel("● Device Standby")
        self.lbl_rec_status.setStyleSheet("font-size: 9px; color: #aaa;")
        
        rec_ctrl.addWidget(self.btn_toggle_rec, 2)
        rec_ctrl.addWidget(self.lbl_rec_status, 1)
        layout.addLayout(rec_ctrl)

        lbl_daftar = QLabel("Daftar Rekaman Mesin (.CSV):")
        lbl_daftar.setStyleSheet("font-size: 9px;")
        layout.addWidget(lbl_daftar)
        
        self.log_list = QListWidget()
        self.log_list.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: white; font-size: 8px;")
        self.log_list.itemDoubleClicked.connect(self._open_log_detail)
        layout.addWidget(self.log_list)

        btn_lay = QHBoxLayout()
        self.btn_watch_log = QPushButton("Buka Panel Deteksi")
        self.btn_export_excel = QPushButton("Export ke Excel")
        self.btn_delete_log = QPushButton("Hapus Rekaman")

        self.btn_watch_log.setStyleSheet("background-color: #cfcfcf; color: #000000; font-size: 8px; height: 20px; font-weight: bold; border-radius: 3px;")
        self.btn_export_excel.setStyleSheet("background-color: #217346; color: #ffffff; font-size: 8px; height: 20px; font-weight: bold; border-radius: 3px;")
        self.btn_delete_log.setStyleSheet("background-color: #cfcfcf; color: #000000; font-size: 8px; height: 20px; font-weight: bold; border-radius: 3px;")

        self.btn_watch_log.clicked.connect(self._open_log_detail_from_button)
        self.btn_export_excel.clicked.connect(self._export_selected_log_to_excel)
        self.btn_delete_log.clicked.connect(self._delete_selected_log)

        btn_lay.addWidget(self.btn_watch_log)
        btn_lay.addWidget(self.btn_export_excel)
        btn_lay.addWidget(self.btn_delete_log)
        layout.addLayout(btn_lay)

        lbl_hint = QLabel("* Klik file rekaman untuk membuka panel forensik anomali.")
        lbl_hint.setStyleSheet("font-size: 9px; color: #888; font-style: italic;")
        layout.addWidget(lbl_hint)

        return page


    def _toggle_recording(self):
        if self.selected_slot_idx < 0:
            QMessageBox.warning(self, "Perhatian", "Silakan pilih Slot Mesin terlebih dahulu!")
            return
            
        if not self.recording:
            filename = os.path.join(LOG_DIR, f"log_{datetime.now().strftime('%d%m%Y_%H%M%S')}.csv")
            try:
                self.csv_file = open(filename, 'w', newline='')
                self.csv_writer = csv.writer(self.csv_file)
                self.csv_writer.writerow([
                    'timestamp', 'machine_type', 'rms_v', 'rms_a', 'temp', 'vib_x', 'vib_y', 'vib_z', 
                    'rpm', 'mahalanobis_d2', 'status', 'health_score', 'trend', 'servis_est', 'ml_label', 'diag_label', 'kurtosis'
                ])
                self.record_start_time = time.perf_counter()
                self.recording = True
                self.btn_toggle_rec.setText("BERHENTI RECORDING")
                self.btn_toggle_rec.setStyleSheet(f"background-color: {COL_BAD}; color: #ffffff; font-weight: bold; font-size: 8px; height: 22px;")
                self.lbl_rec_status.setText(f"● MENULIS -> {os.path.basename(filename)}")
            except Exception as e:
                print(f"Gagal membuat file log: {e}")
        else:
            self.recording = False
            if self.csv_file:
                self.csv_file.close()
                self.csv_file = None
            self.btn_toggle_rec.setText("MULAI RECORDING")
            self.btn_toggle_rec.setStyleSheet("background-color: #cfcfcf; color: #000000; font-weight: bold; font-size: 8px; height: 22px;")
            self.lbl_rec_status.setText("● Berkas Tersimpan")
            self._refresh_log_list()


    def _refresh_log_list(self):
        self.log_list.clear()
        if os.path.isdir(LOG_DIR):
            for file in sorted(os.listdir(LOG_DIR), reverse=True):
                if file.endswith(".csv"):
                    self.log_list.addItem(file)


    def _open_log_detail(self, item):
        if item is None:
            return
        filename = item.text()
        filepath = os.path.join(LOG_DIR, filename)

        self.logdet_timer.stop()
        self.logdet_play_index = 0
        self.btn_logdet_play_pause.setText("▶ PLAY")
        self.lbl_logdet_title.setText(f"HASIL DETEKSI REKAMAN: {filename}")

        (self.logdet_times, self.logdet_v_vals, self.logdet_a_vals,
         self.logdet_t_vals) = self._logdet_load_csv(filepath)

        self.slider_logdet.setMaximum(max(0, len(self.logdet_times) - 1))
        self.slider_logdet.setValue(0)
        if self.logdet_times:
            self._logdet_render_frame(0)
        else:
            self.curve_logdet_v.clear()
            self.curve_logdet_a.clear()
            self.curve_logdet_temp.clear()
            self.lbl_logdet_pos.setText("0 / 0")
        self._logdet_render_diagnosis_summary()

        self._change_page(5)


    def _open_log_detail_from_button(self):
        current_item = self.log_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "Perhatian", "Pilih file rekaman (.csv) terlebih dahulu!")
            return
        self._open_log_detail(current_item)


    def _delete_selected_log(self):
        current_item = self.log_list.currentItem()
        if not current_item:
            return
            
        filename = current_item.text()
        filepath = os.path.join(LOG_DIR, filename)
        
        reply = QMessageBox.question(
            self, 'Konfirmasi Hapus', f"Apakah Anda yakin ingin menghapus berkas rekaman {filename}?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
                self._refresh_log_list()
                self.lbl_rec_status.setText("● BERKAS DIHAPUS")
            except Exception as e:
                print(f"Gagal menghapus file: {e}")


    def _export_selected_log_to_excel(self):
        current_item = self.log_list.currentItem()
        if not current_item:
            QMessageBox.information(self, "Pilih Rekaman Dulu", "Klik salah satu berkas rekaman di daftar dulu.")
            return

        if openpyxl is None:
            QMessageBox.critical(self, "Library Belum Terinstall", "Fitur export ke Excel butuh library 'openpyxl'.")
            return

        filename = current_item.text()
        csv_path = os.path.join(LOG_DIR, filename)
        xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"

        try:
            with open(csv_path, 'r') as f:
                rows = list(csv.reader(f))

            if not rows:
                QMessageBox.warning(self, "Berkas Kosong", "Berkas rekaman ini tidak memiliki data.")
                return

            header, data_rows = rows[0], rows[1:]

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Sensor ESP32"

            header_fill = PatternFill(start_color="1c1e22", end_color="1c1e22", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            ws.append(header)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            status_col_idx = None
            for i, col_name in enumerate(header):
                if col_name.strip().lower() == "status":
                    status_col_idx = i
                    break

            status_fill = {
                "diam":    PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid"),
                "normal":  PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid"), 
                "waspada": PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid"),  
                "bahaya":  PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid"),   
            }

            for row in data_rows:
                converted_row = []
                for value in row:
                    try:
                        converted_row.append(float(value))
                    except (ValueError, TypeError):
                        converted_row.append(value)
                ws.append(converted_row)

                if status_col_idx is not None:
                    status_val = str(row[status_col_idx]).strip().lower()
                    fill = status_fill.get(status_val)
                    if fill:
                        ws.cell(row=ws.max_row, column=status_col_idx + 1).fill = fill

            for col_cells in ws.columns:
                max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
                ws.column_dimensions[col_cells[0].column_letter].width = max_len + 2

            ws.freeze_panes = "A2"  
            wb.save(xlsx_path)

            QMessageBox.information(
                self, "Export Berhasil",
                f"Data berhasil diexport ke:\n{xlsx_path}\n\nTotal {len(data_rows)} baris data."
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Gagal", f"Terjadi kesalahan saat export:\n{e}")

