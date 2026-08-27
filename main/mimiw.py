import sys
import os
import csv
import json
import time
from datetime import datetime
from collections import deque

try:
    import serial
    import serial.tools.list_ports
    import threading
except ImportError:
    serial = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None

from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton, QHBoxLayout, QVBoxLayout,
    QGridLayout, QStackedWidget, QFrame, QListWidget, QListWidgetItem, QComboBox, QMessageBox,
    QSlider, QSizePolicy, QProgressBar, QLineEdit, QDialog, QFormLayout
)
from PyQt5.QtCore import QTimer, Qt, QPoint, QEvent
from PyQt5.QtGui import QFont, QPainter, QLinearGradient, QColor, QPolygon, QPen, QBrush
import pyqtgraph as pg

SERIAL_PORT = '/dev/ttyACM0'
BAUD_RATE = 115200
LOG_DIR = "logs"
CONFIG_FILE = "machines_config.json"
HISTORY_FILE = "check_history.json"
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)

ESP32_USB_HINTS = ["CH340", "CH343", "CP210", "USB-SERIAL", "USB SERIAL", "FTDI", "SILICON LABS", "COM3", "ACM0"]

DATASET_DIR = "Dataset"
KOLOM_DATASET = [
    "waktu_lokal", "rms_v", "rms_x", "rms_y", "rms_z", "rms_a",
    "cur", "cur_raw_adc", "temp", "temp_raw", "rpm", "snr", "d2",
    "status", "e_unbalance", "e_misalign", "e_bpfo", "e_bpfi",
    "diagnosis", "diag_conf",
    "e_audio_low", "e_audio_mid", "e_audio_high",
    "audio_diagnosis", "audio_diag_conf",
    "ml_label", "ml_conf",
    "roughness", "brightness",
    "ground_truth"
]
KONDISI_REKAM_OPSI = [
    ("O", "kondisiNormal"),
    ("U", "kondisiUnbalance"),
    ("M", "kondisiMisalignment"),
    ("F", "kondisiBearingFaulting"),
    ("L", "kondisiLubrication"),
    ("D", "kondisiMati"),
]
SLOT_LABEL_UNTUK_NAMA_FILE = {
    0: "slot0_1400rpm",
    1: "slot1_2800rpm",
}
if not os.path.exists(DATASET_DIR):
    os.makedirs(DATASET_DIR)

COL_BG_MAIN = "#181b20"
COL_PANEL_DARK = "#101216"
COL_ACCENT = "#38bdf8"
COL_ACCENT_DIM = "#1e2733"
COL_TEXT_LIGHT = "#e8eaed"
COL_TEXT_DIM = "#7c8592"
COL_OK = "#34d399"
COL_WARN = "#fbbf24"
COL_BAD = "#f87171"
COL_IDLE = "#94a3b8"
COL_HEADER_BG = "#1c2027"

D2_THRESHOLD_WASPADA = 7.815
D2_THRESHOLD_BAHAYA = 11.345

STATUS_SEVERITY = {"diam": -1, "normal": 0, "waspada": 1, "bahaya": 2}

JENIS_MESIN_OPSI = [
    ("🌀", "Kipas Angin", "Mensirkulasikan/mendinginkan udara di ruangan atau area kerja."),
    ("💧", "Pompa Air", "Memompa air dari satu tempat ke tempat lain (misal: dari sumur/sumber ke tandon penyimpanan)."),
    ("🥤", "Blender", "Menghaluskan atau mencampur bahan makanan/minuman."),
    ("🧹", "Vacuum Cleaner", "Menyedot debu dan kotoran dari lantai atau permukaan lainnya."),
    ("🛠️", "Bor Listrik", "Melubangi atau mengencangkan material pada pekerjaan bengkel/konstruksi."),
    ("🧵", "Mesin Jahit", "Menjahit kain untuk produksi pakaian atau kerajinan tekstil."),
    ("🌬️", "Blower / Exhaust Fan", "Meniupkan/menghisap udara untuk proses industri (misal: sirkulasi udara, pengering, atau tungku UMKM)."),
    ("📦", "Motor Konveyor", "Menggerakkan sabuk konveyor untuk memindahkan barang pada proses produksi."),
    ("⚙️", "Mesin/Motor Rotasi Lainnya", "Mesin rotasi umum yang belum masuk kategori spesifik di atas."),
]

class AddMachineDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Tambah Slot Mesin & Protokol Hardware")
        self.setFixedSize(300, 235)  # FIX (27 Agustus 2026): dikecilin, satu baris field (Kegunaan) dihapus
        self.setStyleSheet(f"background-color: {COL_BG_MAIN}; color: {COL_TEXT_LIGHT}; font-family: Arial; font-size: 8px;")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(4)

        lbl_head = QLabel("KONFIGURASI SLOT MESIN BARU")
        lbl_head.setStyleSheet(f"font-weight: bold; color: {COL_ACCENT}; font-size: 9px;")
        layout.addWidget(lbl_head)

        form_layout = QFormLayout()
        form_layout.setSpacing(4)

        self.cmb_jenis = QComboBox()
        self.cmb_jenis.addItems([f"{icon}  {label}" for icon, label, _kegunaan in JENIS_MESIN_OPSI])
        self.cmb_jenis.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: {COL_TEXT_LIGHT};")
        form_layout.addRow("Jenis Mesin:", self.cmb_jenis)

        self.cmb_slot = QComboBox()
        self.cmb_slot.addItems([str(i) for i in range(10)])
        self.cmb_slot.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: {COL_TEXT_LIGHT};")
        form_layout.addRow("Slot (0-9):", self.cmb_slot)

        self.cmb_bearing = QComboBox()
        self.cmb_bearing.addItems(["Rolling Bearing (B)", "Bushing / Tanpa Bearing (N)"])
        self.cmb_bearing.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: {COL_TEXT_LIGHT};")
        form_layout.addRow("Tipe Bearing:", self.cmb_bearing)

        self.cmb_cluster = QComboBox()
        self.cmb_cluster.addItems(["Klaster A (~1400 RPM - V)", "Klaster B (~2800 RPM - W)"])
        self.cmb_cluster.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: {COL_TEXT_LIGHT};")
        form_layout.addRow("Klaster RPM:", self.cmb_cluster)

        self.cmb_calibration_mode = QComboBox()
        self.cmb_calibration_mode.addItems(["Dengan Kalibrasi (180s Baseline)", "Tanpa Kalibrasi (Direct Monitoring)"])
        self.cmb_calibration_mode.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: {COL_TEXT_LIGHT};")
        form_layout.addRow("Mode Operasi:", self.cmb_calibration_mode)

        layout.addLayout(form_layout)

        btn_layout = QHBoxLayout()
        self.btn_save = QPushButton("Simpan Konfigurasi")
        self.btn_save.setStyleSheet(f"background-color: {COL_ACCENT}; color: #000; font-weight: bold; padding: 4px;")
        self.btn_save.clicked.connect(self.accept)

        self.btn_cancel = QPushButton("Batal")
        self.btn_cancel.setStyleSheet("background-color: #444; color: #000; padding: 4px;")
        self.btn_cancel.clicked.connect(self.reject)

        btn_layout.addWidget(self.btn_save)
        btn_layout.addWidget(self.btn_cancel)
        layout.addLayout(btn_layout)

    def get_data(self):
        is_no_calib = (self.cmb_calibration_mode.currentIndex() == 1)
        jenis_icon, jenis_label, jenis_kegunaan = JENIS_MESIN_OPSI[self.cmb_jenis.currentIndex()]
        return {
            "name": jenis_label,
            "icon": jenis_icon,
            "slot_char": self.cmb_slot.currentText(),
            "bearing_cmd": "B" if self.cmb_bearing.currentIndex() == 0 else "N",
            "cluster_cmd": "V" if self.cmb_cluster.currentIndex() == 0 else "W",
            "no_calib": is_no_calib,
            "rpm_cluster": "Custom RPM",
            "baseline_d2": 9.49,
            "fw_cluster": "Custom",
            "kegunaan": jenis_kegunaan,
        }

class _GaugeBar(QWidget):
    def __init__(self, min_val, max_val):
        super().__init__()
        self.min_val = min_val
        self.max_val = max_val
        self.val = min_val

    def set_value(self, val):
        self.val = max(self.min_val, min(self.max_val, val))
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        w = self.width()
        h = self.height() - 4

        grad = QLinearGradient(0, 0, w, 0)
        grad.setColorAt(0.0, QColor(COL_OK))
        grad.setColorAt(0.5, QColor(COL_WARN))
        grad.setColorAt(1.0, QColor(COL_BAD))

        painter.setBrush(QBrush(grad))
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(0, 0, w, h, 2, 2)

        ratio = (self.val - self.min_val) / (self.max_val - self.min_val) if self.max_val > self.min_val else 0
        px = int(ratio * w)

        poly = QPolygon([
            QPoint(px, h),
            QPoint(max(0, px - 4), h + 4),
            QPoint(min(w, px + 4), h + 4)
        ])
        painter.setBrush(QBrush(QColor("#ffffff")))
        painter.setPen(QPen(QColor("#ffffff"), 1))
        painter.drawPolygon(poly)

class GradientGauge(QWidget):
    def __init__(self, title, min_val, max_val, t_warn, t_danger, unit=""):
        super().__init__()
        self.min_val = min_val
        self.max_val = max_val
        self.t_warn = t_warn
        self.t_danger = t_danger
        self.unit = unit
        self.current_val = min_val

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(1)

        self.lbl_title = QLabel(title)
        self.lbl_title.setStyleSheet(f"font-size: 11px; font-weight: bold; color: {COL_TEXT_LIGHT};")
        layout.addWidget(self.lbl_title)

        self.bar = _GaugeBar(min_val, max_val)
        self.bar.setFixedHeight(18)
        layout.addWidget(self.bar)

        self.lbl_status = QLabel("Normal")
        self.lbl_status.setStyleSheet(f"font-size: 10px; color: {COL_TEXT_DIM};")
        self.lbl_status.setWordWrap(True)
        layout.addWidget(self.lbl_status)

    def set_value(self, val, status_key="normal"):
        if val is None: return
        self.current_val = val
        self.bar.set_value(val)

        if status_key == "diam":
            status, desc = "Diam / Off", "Mati"
            col = COL_IDLE
        elif val >= self.t_danger:
            status, desc = "Extreme", "Bahaya"
            col = COL_BAD
        elif val >= self.t_warn:
            status, desc = "Moderate", "Waspada"
            col = COL_WARN
        else:
            status, desc = "Normal", "Aman"
            col = COL_OK

        self.lbl_status.setText(f"<span style='color:{col}; font-weight:bold;'>{status}</span>")

class Dashboard(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("HMI | VIBRIS PIMNAS - 1800+ ENTERPRISE ENGINE (PERSISTENT STORAGE)")
        self.setWindowFlags(
            Qt.Window | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint
        )
        self.setFixedSize(480, 320)
        self.setStyleSheet(f"background-color: {COL_BG_MAIN}; color: {COL_TEXT_LIGHT}; font-family: Arial;")

        screen = QApplication.primaryScreen()
        if screen is not None:
            screen_geo = screen.geometry()
            self.move(screen_geo.x(), screen_geo.y())

        self.data_len = 50
        self.time_buffer = deque(maxlen=self.data_len)
        self.v_buffer = deque(maxlen=self.data_len)
        self.vx_buffer = deque(maxlen=self.data_len)
        self.vy_buffer = deque(maxlen=self.data_len)
        self.vz_buffer = deque(maxlen=self.data_len)
        self.a_buffer = deque(maxlen=self.data_len)
        self.temp_buffer = deque(maxlen=self.data_len)
        self.rpm_buffer = deque(maxlen=self.data_len)
        self.d2_buffer = deque(maxlen=self.data_len)
        
        self.history_window_size = 100
        self.hist_v_buf = deque(maxlen=self.history_window_size)
        self.hist_a_buf = deque(maxlen=self.history_window_size)
        self.hist_temp_buf = deque(maxlen=self.history_window_size)
        self.hist_vx_buf = deque(maxlen=self.history_window_size)
        self.hist_vy_buf = deque(maxlen=self.history_window_size)
        self.hist_vz_buf = deque(maxlen=self.history_window_size)
        self.hist_rpm_buf = deque(maxlen=self.history_window_size)
        self.hist_d2_buf = deque(maxlen=self.history_window_size)

        self.tick = 0
        self.last_processed_tick = 0
        self.last_raw_line = ""
        self.last_packet_time = time.time()
        self.packet_loss_flag = False
        self.last_auto_snapshot_status = "normal"

        self.fft_hz_buffer = []
        self.fft_mag_buffer = []

        self.current_health_score = 0.0
        self.current_trend = "Menunggu"
        self.current_servis = "Menunggu"
        self.current_ml_label = "N/A"
        self.current_ml_conf = 0.0
        self.current_diag_label = "N/A"
        self.current_diag_conf = 0.0
        self.current_kurtosis = 0.0
        self.current_diagnosis_flags = "N/A"

        self.logdet_times = []
        self.logdet_v_vals = []
        self.logdet_a_vals = []
        self.logdet_t_vals = []
        self.logdet_play_index = 0
        self.logdet_play_speed = 1.0
        self.logdet_timer = QTimer(self)
        self.logdet_timer.timeout.connect(self._logdet_step)

        self.calibration_timer = QTimer(self)
        self.calibration_timer.timeout.connect(self._calibration_countdown_tick)
        self.calibration_time_left = 180
        self.calibrating_machine_idx = -1

        self.delay_calibration_timer = QTimer(self)
        self.delay_calibration_timer.setSingleShot(True)
        self.delay_calibration_timer.timeout.connect(self._execute_actual_calibration)
        self.pending_calibration_idx = -1
        self.delay_seconds_left = 4

        self.check_session_timer = QTimer(self)
        self.check_session_timer.timeout.connect(self._check_session_countdown_tick)
        self.check_session_time_left = 60
        self.is_check_session_active = False

        self.current_v = None
        self.current_vx = None
        self.current_vy = None
        self.current_vz = None
        self.current_a = None
        self.current_temp = None
        self.current_rpm = None
        self.current_d2 = None
        self.current_status_device = ""

        self.session_sample_count = 0
        self.session_rpm_sum = 0.0
        self.session_d2_max = 0.0
        self.session_worst_status = "Normal"
        self.session_waspada_count = 0
        self.session_bahaya_count = 0
        self.anomaly_events = []

        self.recording = False
        self.csv_file = None
        self.csv_writer = None
        self.ser = None
        self.serial_connected = False
        self.last_raw_data = {}

        self.dataset_recording = False
        self.dataset_csv_file = None
        self.dataset_csv_writer = None
        self.dataset_summary_path = None

        self.machines = self._load_machines_config()
        self.check_history = self._load_check_history()
        self.selected_machine_idx = 0 if self.machines else -1
        self.machine_delete_mode = False
        self.is_technician_mode = True

        root = QVBoxLayout(self)
        root.setContentsMargins(2, 2, 2, 2)
        root.setSpacing(2)

        header = QHBoxLayout()
        header.setSpacing(2)
        header.setContentsMargins(2, 2, 2, 2)
        self.header_frame = QFrame()
        self.header_frame.setStyleSheet(
            f"background-color: {COL_HEADER_BG}; border-radius: 3px; "
            f"border-bottom: 2px solid {COL_ACCENT};"
        )
        self.header_frame.setLayout(header)

        initial_slot_str = self.machines[0]["slot_char"] if self.machines else "-"
        initial_name_str = self.machines[0]["name"] if self.machines else "- Pilih Mesin -"
        self.lbl_machine_active = QPushButton(f"⚙️ [{initial_slot_str}] {initial_name_str}")
        self.lbl_machine_active.setStyleSheet(
            f"QPushButton {{ background-color: {COL_ACCENT_DIM}; color: {COL_TEXT_LIGHT}; font-size: 8px; "
            f"font-weight: bold; padding: 3px 4px; border: 1px solid #2a3542; border-radius: 2px; text-align: left; }}"
            f"QPushButton:hover {{ border: 1px solid {COL_ACCENT}; }}"
        )
        self.lbl_machine_active.clicked.connect(lambda: self._change_page(2 if self.is_technician_mode else 1))
        header.addWidget(self.lbl_machine_active, 2)

        def _header_btn_style(border_color, text_color):
            return (
                f"QPushButton {{ background-color: transparent; color: {text_color}; "
                f"font-weight: bold; font-size: 7px; padding: 3px 3px; "
                f"border: 1px solid {border_color}; border-radius: 2px; }}"
                f"QPushButton:hover {{ background-color: {border_color}; color: #101216; }}"
            )

        self.top_action_buttons_container = QWidget()
        tac_layout = QHBoxLayout(self.top_action_buttons_container)
        tac_layout.setContentsMargins(0, 0, 0, 0)
        tac_layout.setSpacing(1)

        self.btn_cek60s = QPushButton("CEK(K)")
        self.btn_cek60s.setStyleSheet(_header_btn_style(COL_OK, COL_OK))
        self.btn_cek60s.clicked.connect(self._trigger_check_session)
        tac_layout.addWidget(self.btn_cek60s)

        self.btn_slot_summary = QPushButton("RINGKAS(P)")
        self.btn_slot_summary.setStyleSheet(_header_btn_style(COL_ACCENT, COL_ACCENT))
        self.btn_slot_summary.clicked.connect(lambda: self._send_command('P'))
        tac_layout.addWidget(self.btn_slot_summary)

        self.btn_del_base = QPushButton("HAPUS(Z)")
        self.btn_del_base.setStyleSheet(_header_btn_style(COL_WARN, COL_WARN))
        self.btn_del_base.clicked.connect(lambda: self._send_command('Z'))
        tac_layout.addWidget(self.btn_del_base)

        self.btn_recal = QPushButton("KALIB(R)")
        self.btn_recal.setStyleSheet(_header_btn_style(COL_ACCENT, COL_ACCENT))
        self.btn_recal.clicked.connect(self._trigger_recalibration)
        tac_layout.addWidget(self.btn_recal)

        self.btn_reboot_esp = QPushButton("REBOOT(X)")
        self.btn_reboot_esp.setStyleSheet(_header_btn_style(COL_BAD, COL_BAD))
        self.btn_reboot_esp.clicked.connect(self._confirm_and_reboot)
        tac_layout.addWidget(self.btn_reboot_esp)

        self.btn_debug = QPushButton("DBG")
        self.btn_debug.setStyleSheet(_header_btn_style(COL_TEXT_DIM, COL_TEXT_DIM))
        self.btn_debug.clicked.connect(self._show_debug_info)
        tac_layout.addWidget(self.btn_debug)

        header.addWidget(self.top_action_buttons_container)

        self.btn_mode_toggle = QPushButton("MODE: ENG")
        self.btn_mode_toggle.setStyleSheet(
            "QPushButton { background-color: #fbbf24; color: #101216; font-weight: bold; font-size: 7px; padding: 3px; border-radius: 2px; border: 1px solid #fbbf24; }"
            "QPushButton:hover { background-color: #f59e0b; }"
        )
        self.btn_mode_toggle.clicked.connect(self._toggle_mode_system)
        header.addWidget(self.btn_mode_toggle)

        self.lbl_conn_dot = QLabel("●")
        self.lbl_conn_dot.setStyleSheet(f"font-size: 10px; font-weight: bold; color: {COL_BAD};")
        header.addWidget(self.lbl_conn_dot)

        time_box = QVBoxLayout()
        time_box.setSpacing(0)
        time_box.setContentsMargins(0, 0, 0, 0)

        self.time_lbl = QLabel("--:--:--")
        self.time_lbl.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {COL_OK};")
        self.time_lbl.setAlignment(Qt.AlignRight)
        time_box.addWidget(self.time_lbl)

        self.date_lbl = QLabel("--/--/----")
        self.date_lbl.setStyleSheet(f"font-size: 6px; color: {COL_OK};")
        self.date_lbl.setAlignment(Qt.AlignRight)
        time_box.addWidget(self.date_lbl)

        header.addLayout(time_box)
        root.addWidget(self.header_frame)

        self.stack = QStackedWidget()

        self.stack.addWidget(self._page_awam_beranda())        
        self.stack.addWidget(self._page_awam_mesin_saya())     

        self.stack.addWidget(self._page_machine_select())     
        self.stack.addWidget(self._page_log_detail())          
        self.stack.addWidget(self._page_raw())                 
        self.stack.addWidget(self._page_recording())           
        self.stack.addWidget(self._page_processed())           
        self.stack.addWidget(self._page_summary())             
        self.stack.addWidget(self._page_awam_riwayat())        
        self.stack.addWidget(self._page_awam_detail_mesin())   
        self.stack.addWidget(self._page_awam_proses())         
        self.stack.addWidget(self._page_awam_cek_status())     

        root.addWidget(self.stack, 1)

        self.nav_bottom_widget = QWidget()
        nav_bottom = QHBoxLayout(self.nav_bottom_widget)
        nav_bottom.setContentsMargins(0, 0, 0, 0)
        nav_bottom.setSpacing(2)
        
        self.btn_nav1 = QPushButton("BERANDA")
        self.btn_nav2 = QPushButton("MESIN SAYA")
        self.btn_nav3 = QPushButton("CEK STATUS")
        self.btn_nav4 = QPushButton("REBOOT ESP")

        self.menu_buttons = [self.btn_nav1, self.btn_nav2, self.btn_nav3, self.btn_nav4]

        for i, btn in enumerate(self.menu_buttons):
            btn.setFixedHeight(30)  
            btn.setStyleSheet(self._menu_style(False))
            btn.clicked.connect(lambda checked, idx=i: self._change_page_by_nav(idx))
            nav_bottom.addWidget(btn)
        
        root.addWidget(self.nav_bottom_widget)

        self._init_serial_connection()

        self.main_timer = QTimer(self)
        self.main_timer.timeout.connect(self._update_gui)
        self.main_timer.start(200) 

        self._apply_mode_ui_layout()
        self._refresh_log_list()
        self.show()

    def _load_machines_config(self):
        default_machines = [
            {"name": "Blower Industri UMKM", "icon": "🌀", "slot_char": "0", "bearing_cmd": "B", "cluster_cmd": "W", "no_calib": True, "rpm_cluster": "Medium (2800 RPM)", "baseline_d2": 9.49, "fw_cluster": "Klaster B",
             "kegunaan": "Meniupkan/menghisap udara untuk proses industri (misal: sirkulasi udara, pengering, atau tungku UMKM)."},
            {"name": "Motor Induksi Pompa Air", "icon": "💧", "slot_char": "1", "bearing_cmd": "A", "cluster_cmd": "V", "no_calib": True, "rpm_cluster": "Low (1400 RPM)", "baseline_d2": 8.00, "fw_cluster": "Klaster A",
             "kegunaan": "Memompa air dari satu tempat ke tempat lain (misal: dari sumur/sumber ke tandon penyimpanan)."},
        ]
        machines = default_machines
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, list) and len(data) > 0:
                        machines = data
            except Exception as e:
                print(f"Gagal memuat konfigurasi mesin: {e}")

        healed = False
        for i in range(min(2, len(machines))):
            if not machines[i].get("no_calib", False):
                machines[i]["no_calib"] = True
                healed = True
        if healed:
            self.machines = machines
            self._save_machines_config()

        return machines

    def _save_machines_config(self):
        try:
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.machines, f, indent=4)
        except Exception as e:
            print(f"Gagal menyimpan konfigurasi mesin: {e}")

    def _load_check_history(self):
        if os.path.exists(HISTORY_FILE):
            try:
                with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        return data
            except Exception as e:
                print(f"Gagal memuat riwayat cek: {e}")
        return []

    def _save_check_history(self):
        try:
            with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.check_history[-100:], f, indent=2)
        except Exception as e:
            print(f"Gagal menyimpan riwayat cek: {e}")

    def _record_check_history(self, status_text, health_score):
        machine_name = (self.machines[self.selected_machine_idx]["name"]
                         if 0 <= self.selected_machine_idx < len(self.machines) else "Tidak diketahui")
        entry = {
            "waktu": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "mesin": machine_name,
            "status": status_text,
            "health_score": health_score,
        }
        self.check_history.append(entry)
        self._save_check_history()
        if hasattr(self, 'riwayat_list'):
            self._rebuild_riwayat_list()
        self._refresh_wajah_alat()

    def closeEvent(self, event):
        self._save_machines_config()
        if self.csv_file:
            self.csv_file.close()
        if self.dataset_csv_file:
            self.dataset_csv_file.close()
        event.accept()

    def _toggle_mode_system(self):
        self.is_technician_mode = not self.is_technician_mode
        self._apply_mode_ui_layout()

    def _apply_mode_ui_layout(self):
        if self.is_technician_mode:
            self.btn_mode_toggle.setText("MODE: ENG")
            self.btn_mode_toggle.setStyleSheet(
                "QPushButton { background-color: #fbbf24; color: #101216; font-weight: bold; font-size: 7px; padding: 3px; border-radius: 2px; border: 1px solid #fbbf24; }"
                "QPushButton:hover { background-color: #f59e0b; }"
            )
            self.top_action_buttons_container.show()
            self.btn_nav4.show()

            self.btn_nav1.setText("RAW READING")
            self.btn_nav2.setText("LOGS SAVES")
            self.btn_nav3.setText("PROCESSED (FFT)")
            self.btn_nav4.setText("SUMMARY")

            self._change_page(4)
        else:
            self.btn_mode_toggle.setText("MODE: USER")
            self.btn_mode_toggle.setStyleSheet(
                "QPushButton { background-color: #38bdf8; color: #101216; font-weight: bold; font-size: 7px; padding: 3px; border-radius: 2px; border: 1px solid #38bdf8; }"
                "QPushButton:hover { background-color: #0ea5e9; }"
            )
            self.top_action_buttons_container.hide()

            self.btn_nav1.setText("BERANDA")
            self.btn_nav2.setText("MESIN SAYA")
            self.btn_nav3.setText("CEK STATUS")
            self.btn_nav4.hide()

            self._refresh_wajah_alat()
            self._change_page(0)

    def _menu_style(self, active):
        if active:
            return f"background-color: {COL_ACCENT}; color: #000000; font-size: 9px; font-weight: bold; border: 1px solid white; border-radius: 3px;"
        return f"background-color: #cfcfcf; color: #000000; font-size: 9px; font-weight: bold; border: 1px solid #444; border-radius: 3px;"

    def _change_page(self, idx):
        self.stack.setCurrentIndex(idx)

    def _sesi_sedang_aktif(self):
        return (
            self.is_check_session_active
            or self.calibration_timer.isActive()
            or self.delay_calibration_timer.isActive()
        )

    def _change_page_by_nav(self, nav_idx):
        if self.is_technician_mode:
            stack_mapping = {0: 4, 1: 5, 2: 6, 3: 7}
            target_stack = stack_mapping.get(nav_idx, 4)
            self._change_page(target_stack)
            for i, btn in enumerate(self.menu_buttons):
                btn.setStyleSheet(self._menu_style(i == nav_idx))
        elif self._sesi_sedang_aktif():
            self._change_page(10)
        else:
            if nav_idx == 0:
                self._refresh_wajah_alat()
                self._change_page(0)
            elif nav_idx == 1:
                self._change_page(1)
            elif nav_idx == 2:
                self._change_page(11)

            for i, btn in enumerate(self.menu_buttons):
                btn.setStyleSheet(self._menu_style(i == nav_idx if nav_idx < 3 else False))

    def _confirm_and_reboot(self):
        sedang_sibuk = (
            self.is_check_session_active
            or self.calibration_timer.isActive()
            or self.delay_calibration_timer.isActive()
        )
        if sedang_sibuk:
            pesan = ("Ada sesi Cek/Kalibrasi yang SEDANG BERJALAN.\n\n"
                     "Reboot sekarang akan MEMOTONG sesi itu sebelum selesai.\n\n"
                     "Yakin tetap reboot ESP32?")
        else:
            pesan = "Yakin mau reboot ESP32? Koneksi akan terputus sesaat."

        jawaban = QMessageBox.question(
            self, "Konfirmasi Reboot ESP32", pesan,
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if jawaban == QMessageBox.Yes:
            self._send_command('X')

    def _trigger_recalibration(self):
        if self.is_check_session_active:
            QMessageBox.warning(self, "Ditolak", "Kalibrasi ditolak! Sesi check (K) sedang berjalan.")
            return
        current_m = self.machines[self.selected_machine_idx] if self.selected_machine_idx >= 0 else None
        if current_m and current_m.get("no_calib", False):
            QMessageBox.information(self, "Mode Tanpa Kalibrasi", "Slot ini dikonfigurasi dalam Mode Tanpa Kalibrasi (Direct Monitoring / Preset Pabrik Terkunci). Kalibrasi dilewati.")
            return
        self._start_delayed_calibration(self.selected_machine_idx)

    def _trigger_check_session(self):
        if self.is_check_session_active:
            QMessageBox.information(self, "Sesi Sedang Berjalan", "Sesi Cek masih berjalan, tunggu sampai selesai dulu.")
            if not self.is_technician_mode:
                self._change_page(10)
            return
        if self.calibration_timer.isActive() or self.delay_calibration_timer.isActive():
            QMessageBox.warning(self, "Ditolak", "Sesi check ditolak! Perangkat masih dalam fase persiapan/kalibrasi (R).")
            return
        self._send_command('K')
        if self.is_technician_mode:
            self._start_60s_check_session()
        else:
            self._show_proses_loading("MEMPERSIAPKAN CEK", self._start_60s_check_session)

    def _start_60s_check_session(self):
        self.is_check_session_active = True
        self.check_session_time_left = 60
        self.check_session_timer.start(1000)
        self._set_proses_labels("SESI CEK BERJALAN", "60", "Mengumpulkan data analisis operasional...")

    def _check_session_countdown_tick(self):
        if self.check_session_time_left > 0:
            self.check_session_time_left -= 1
            self._set_proses_labels("SESI CEK BERJALAN", str(self.check_session_time_left), "Mengumpulkan data analisis operasional...")
        else:
            self.check_session_timer.stop()
            self.is_check_session_active = False

            status_key = (self.current_status_device or "").strip().lower()
            if status_key == "bahaya":
                self.lbl_beranda_icon.setText("😨")
                self.lbl_beranda_icon.setStyleSheet(f"font-size: 46px; color: {COL_BAD}; border: none;")
                status_tampil = "BAHAYA / RISIKO KERUSAKAN"
                warna_status = COL_BAD
            elif status_key == "waspada":
                self.lbl_beranda_icon.setText("😟")
                self.lbl_beranda_icon.setStyleSheet(f"font-size: 46px; color: {COL_WARN}; border: none;")
                status_tampil = "WASPADA DEVIASI OPERASIONAL"
                warna_status = COL_WARN
            elif status_key == "diam":
                self.lbl_beranda_icon.setText("😴")
                self.lbl_beranda_icon.setStyleSheet(f"font-size: 46px; color: {COL_IDLE}; border: none;")
                status_tampil = "MOTOR SEDANG DIAM / MATI"
                warna_status = COL_IDLE
            else:
                self.lbl_beranda_icon.setText("😊")
                self.lbl_beranda_icon.setStyleSheet(f"font-size: 46px; color: {COL_OK}; border: none;")
                status_tampil = "MESIN AMAN / NORMAL"
                warna_status = COL_OK

            self.lbl_check_text.setStyleSheet(f"font-size: 10px; font-weight: bold; color: {warna_status}; border: none;")
            self.lbl_check_text.setText(status_tampil)

            self.lbl_rincian_getaran.setText(f"{self.current_v:.2f} G")
            self.lbl_rincian_suara.setText(f"{self.current_a:.1f} dB")
            self.lbl_rincian_suhu.setText(f"{self.current_temp:.1f} °C")
            self.lbl_rincian_waktu.setText(datetime.now().strftime("%H:%M:%S"))
            _peta_diagnosis_awam = {
                "UNBALANCE":     "Ketidakseimbangan (Unbalance)",
                "MISALIGNMENT":  "Poros Tidak Lurus (Misalignment)",
                "BEARING_FAULT": "Kerusakan Bearing (Bearing Fault)",
                "BEARING_BPFO":  "Kerusakan Bearing (Bearing Fault)",
                "BEARING_BPFI":  "Kerusakan Bearing (Bearing Fault)",
                "NORMAL":        "Tidak ada",
                "N/A":           "Tidak ada",
            }
            diag_mentah = (self.current_diag_label or "N/A").strip().upper()
            self.lbl_rincian_diagnosis.setText(_peta_diagnosis_awam.get(diag_mentah, diag_mentah))
            self.lbl_desc_text.setText(f"Hasil dari sesi Cek jam {datetime.now().strftime('%H:%M:%S')}")
            self.lbl_estimasi_servis.setText(f"Estimasi servis berikutnya: {self.current_servis}")

            if not self.is_technician_mode:
                self._change_page(11)

            self._record_check_history(status_key or "normal", self.current_health_score)

    def _page_awam_beranda(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(6, 6, 6, 6)
        lay.setSpacing(6)

        box_top = QFrame()
        box_top.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px solid #2a3542; border-radius: 6px;")
        top_lay = QVBoxLayout(box_top)
        top_lay.setAlignment(Qt.AlignCenter)
        top_lay.setContentsMargins(6, 10, 6, 10)

        self.lbl_wajah_icon = QLabel("⚙️")
        self.lbl_wajah_icon.setAlignment(Qt.AlignCenter)
        self.lbl_wajah_icon.setStyleSheet(f"font-size: 40px; color: {COL_ACCENT}; border: none;")
        top_lay.addWidget(self.lbl_wajah_icon)

        self.lbl_wajah_nama = QLabel("- Belum Ada Mesin Dipilih -")
        self.lbl_wajah_nama.setAlignment(Qt.AlignCenter)
        self.lbl_wajah_nama.setWordWrap(True)
        self.lbl_wajah_nama.setStyleSheet(f"font-size: 13px; font-weight: bold; color: {COL_TEXT_LIGHT}; border: none;")
        top_lay.addWidget(self.lbl_wajah_nama)

        self.lbl_wajah_info = QLabel("Pilih mesin di menu MESIN SAYA")
        self.lbl_wajah_info.setAlignment(Qt.AlignCenter)
        self.lbl_wajah_info.setStyleSheet(f"font-size: 9px; color: {COL_TEXT_DIM}; border: none;")
        top_lay.addWidget(self.lbl_wajah_info)
        lay.addWidget(box_top, 2)

        box_edu = QFrame()
        box_edu.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px solid #2a3542; border-radius: 6px;")
        edu_lay = QVBoxLayout(box_edu)
        edu_lay.setContentsMargins(8, 6, 8, 6)
        edu_lay.setSpacing(2)

        lbl_edu_title = QLabel("ℹ️ Tentang Mesin Ini")
        lbl_edu_title.setStyleSheet(f"font-size: 8px; font-weight: bold; color: {COL_ACCENT}; border: none;")
        edu_lay.addWidget(lbl_edu_title)

        self.lbl_wajah_kegunaan = QLabel("-")
        self.lbl_wajah_kegunaan.setWordWrap(True)
        self.lbl_wajah_kegunaan.setStyleSheet(f"font-size: 9px; color: {COL_TEXT_LIGHT}; border: none;")
        edu_lay.addWidget(self.lbl_wajah_kegunaan)

        self.lbl_wajah_rpm = QLabel("-")
        self.lbl_wajah_rpm.setStyleSheet(f"font-size: 8px; color: {COL_TEXT_DIM}; border: none;")
        edu_lay.addWidget(self.lbl_wajah_rpm)
        lay.addWidget(box_edu, 1)

        box_last = QFrame()
        box_last.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px solid #2a3542; border-radius: 6px;")
        last_lay = QVBoxLayout(box_last)
        last_lay.setContentsMargins(8, 6, 8, 6)
        last_lay.setSpacing(2)

        lbl_last_title = QLabel("🕐 Hasil Cek Terakhir")
        lbl_last_title.setStyleSheet(f"font-size: 8px; font-weight: bold; color: {COL_ACCENT}; border: none;")
        last_lay.addWidget(lbl_last_title)

        self.lbl_wajah_last_cek = QLabel("Belum pernah dicek.")
        self.lbl_wajah_last_cek.setWordWrap(True)
        self.lbl_wajah_last_cek.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {COL_TEXT_LIGHT}; border: none;")
        last_lay.addWidget(self.lbl_wajah_last_cek)
        lay.addWidget(box_last, 1)

        btn_riwayat = QPushButton("🕐 Lihat Riwayat Lengkap")
        btn_riwayat.setFixedHeight(28)
        btn_riwayat.setStyleSheet(
            f"QPushButton {{ background-color: {COL_PANEL_DARK}; color: {COL_TEXT_LIGHT}; "
            f"border: 1px solid #2a3542; border-radius: 4px; font-size: 9px; }}"
            f"QPushButton:hover {{ background-color: {COL_ACCENT_DIM}; }}"
        )
        btn_riwayat.clicked.connect(lambda: (self._rebuild_riwayat_list(), self._change_page(8)))
        lay.addWidget(btn_riwayat)

        self._refresh_wajah_alat()
        return page

    def _refresh_wajah_alat(self):
        if not hasattr(self, 'lbl_wajah_nama'):
            return
        if self.selected_machine_idx < 0 or self.selected_machine_idx >= len(self.machines):
            self.lbl_wajah_icon.setText("⚙️")
            self.lbl_wajah_nama.setText("- Belum Ada Mesin Dipilih -")
            self.lbl_wajah_info.setText("Pilih mesin di menu MESIN SAYA")
            self.lbl_wajah_kegunaan.setText("-")
            self.lbl_wajah_rpm.setText("-")
        else:
            m = self.machines[self.selected_machine_idx]
            self.lbl_wajah_icon.setText(m.get("icon", "⚙️"))
            self.lbl_wajah_nama.setText(m["name"])
            mode_str = "Preset Pabrik (Terkunci)" if m.get("no_calib", False) else "Dengan Kalibrasi"
            self.lbl_wajah_info.setText(f"Slot [{m['slot_char']}] — {mode_str}")
            self.lbl_wajah_kegunaan.setText(m.get("kegunaan", "Belum ada deskripsi kegunaan mesin ini."))
            self.lbl_wajah_rpm.setText(f"Kecepatan kerja normal: {m.get('rpm_cluster', '-')}")

        if self.check_history:
            entry = self.check_history[-1]
            status_lower = (entry.get("status") or "").strip().lower()
            emoji_map = {"bahaya": "😨", "waspada": "😟", "diam": "😴"}
            emoji = emoji_map.get(status_lower, "😊")
            self.lbl_wajah_last_cek.setText(
                f"{emoji} {entry.get('mesin','-')}: {entry.get('status','-').upper()} "
                f"({entry.get('waktu','-')})"
            )
        else:
            self.lbl_wajah_last_cek.setText("Belum pernah dicek. Buka Mesin Saya lalu tekan Cek Sekarang.")

    def _page_awam_riwayat(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(2, 2, 2, 2)
        lay.setSpacing(4)

        lbl_title = QLabel("Riwayat Hasil Cek")
        lbl_title.setAlignment(Qt.AlignCenter)
        lbl_title.setStyleSheet(f"font-size: 10px; font-weight: bold; color: {COL_TEXT_LIGHT};")
        lay.addWidget(lbl_title)

        self.riwayat_list = QListWidget()
        self.riwayat_list.setStyleSheet(
            f"QListWidget {{ background-color: {COL_PANEL_DARK}; color: {COL_TEXT_LIGHT}; "
            f"font-size: 8px; border: 1px solid #2a3542; border-radius: 4px; }}"
            f"QListWidget::item {{ padding: 4px; border-bottom: 1px solid #23262c; }}"
        )
        lay.addWidget(self.riwayat_list, 1)

        btn_kembali = QPushButton("← Kembali ke Beranda")
        btn_kembali.setFixedHeight(26)
        btn_kembali.setStyleSheet(
            f"QPushButton {{ background-color: {COL_PANEL_DARK}; color: {COL_TEXT_DIM}; "
            f"border: 1px solid #2a3542; border-radius: 4px; font-size: 8px; }}"
            f"QPushButton:hover {{ background-color: {COL_ACCENT_DIM}; }}"
        )
        btn_kembali.clicked.connect(lambda: self._change_page(0))
        lay.addWidget(btn_kembali)

        self._rebuild_riwayat_list()
        return page

    def _rebuild_riwayat_list(self):
        self.riwayat_list.clear()
        if not self.check_history:
            self.riwayat_list.addItem("Belum ada riwayat Cek. Pilih mesin di Mesin Saya lalu tekan Cek Sekarang.")
            return
        color_map = {"bahaya": COL_BAD, "waspada": COL_WARN, "diam": COL_IDLE}
        for entry in reversed(self.check_history):
            status_lower = (entry.get("status") or "").strip().lower()
            warna = color_map.get(status_lower, COL_OK)
            teks = f"{entry.get('waktu','-')}  |  {entry.get('mesin','-')}  |  {entry.get('status','-').upper()}"
            item = QListWidgetItem(teks)
            item.setForeground(QColor(warna))
            self.riwayat_list.addItem(item)

    def _page_awam_mesin_saya(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 2, 0, 2)
        lay.setSpacing(6)

        lbl_title = QLabel("Pilih Slot Mesin (0-9)")
        lbl_title.setAlignment(Qt.AlignCenter)
        lbl_title.setStyleSheet(f"font-size: 10px; font-weight: bold; color: {COL_TEXT_LIGHT};")
        lay.addWidget(lbl_title)

        self.awam_machine_grid = QGridLayout()
        self.awam_machine_grid.setSpacing(4)
        lay.addLayout(self.awam_machine_grid, 1)

        btn_tambah = QPushButton("➕ Tambah Mesin")
        btn_tambah.setFixedHeight(28)
        btn_tambah.setStyleSheet(
            f"QPushButton {{ background-color: {COL_PANEL_DARK}; color: {COL_TEXT_LIGHT}; "
            f"border: 1px dashed {COL_ACCENT}; border-radius: 4px; font-size: 9px; }}"
            f"QPushButton:hover {{ background-color: {COL_ACCENT_DIM}; }}"
        )
        btn_tambah.clicked.connect(self._open_add_machine_dialog)
        lay.addWidget(btn_tambah)

        self._rebuild_awam_machine_grid()
        return page

    def _rebuild_awam_machine_grid(self):
        while self.awam_machine_grid.count():
            item = self.awam_machine_grid.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

        for i in range(len(self.machines)):
            m = self.machines[i]
            btn = QPushButton()
            btn.setFixedHeight(50)
            
            blay = QVBoxLayout(btn)
            blay.setContentsMargins(2, 2, 2, 2)
            blay.setSpacing(1)

            lbl_ico = QLabel(m["icon"])
            lbl_ico.setAlignment(Qt.AlignCenter)
            lbl_ico.setStyleSheet("font-size: 14px; background: transparent; border: none;")

            lbl_nm = QLabel(m["name"])
            lbl_nm.setAlignment(Qt.AlignCenter)
            lbl_nm.setStyleSheet(f"font-size: 8px; font-weight: bold; color: {COL_TEXT_LIGHT}; background: transparent; border: none;")

            mode_str = "Tanpa Kalibrasi" if m.get("no_calib", False) else "Dengan Kalibrasi"
            inf_text = f"Slot [{m['slot_char']}] — {mode_str}"
            lbl_inf = QLabel(inf_text)
            lbl_inf.setAlignment(Qt.AlignCenter)
            lbl_inf.setStyleSheet(f"font-size: 6px; color: {COL_OK if i < 2 else COL_WARN}; background: transparent; border: none;")

            blay.addWidget(lbl_ico)
            blay.addWidget(lbl_nm)
            blay.addWidget(lbl_inf)

            selected = (i == self.selected_machine_idx)
            border_col = COL_ACCENT if selected else "#2a3542"
            bg = COL_ACCENT_DIM if selected else COL_PANEL_DARK
            btn.setStyleSheet(f"QPushButton {{ background-color: {bg}; border: 1px solid {border_col}; border-radius: 4px; }}")

            btn.clicked.connect(lambda checked, idx=i: self._open_machine_detail(idx))
            self.awam_machine_grid.addWidget(btn, i // 2, i % 2)

    def _select_machine_slot(self, idx, auto_calibrate=True):
        if self.selected_machine_idx == idx and (self.calibration_timer.isActive() or self.delay_calibration_timer.isActive()):
            return

        self.selected_machine_idx = idx
        m = self.machines[idx]

        self.lbl_machine_active.setText(f"⚙️ [{m['slot_char']}] {m['name']}")
        self._refresh_wajah_alat()

        self._send_command(m["slot_char"])
        time.sleep(0.02)
        self._send_command(m["bearing_cmd"])
        time.sleep(0.02)
        self._send_command(m["cluster_cmd"])

        if idx < 2 or m.get("no_calib", False):
            if self.calibration_timer.isActive():
                self.calibration_timer.stop()
            if self.delay_calibration_timer.isActive():
                self.delay_calibration_timer.stop()
            self.calibrating_machine_idx = -1
            self.lbl_check_text.setStyleSheet(f"font-size: 10px; font-weight: bold; color: {COL_OK}; border: none;")
            self.lbl_check_text.setText(f"SLOT {m['slot_char']} SIAP" if idx < 2 else f"SLOT {m['slot_char']} (TANPA KALIBRASI)")
            self.lbl_desc_text.setText("Parameter preset aktif. Tekan CEK STATUS untuk memulai analisis.")
            if not self.is_technician_mode:
                self._change_page(0)
        elif auto_calibrate:
            self._start_delayed_calibration(idx)
        else:
            self.lbl_check_text.setStyleSheet(f"font-size: 10px; font-weight: bold; color: {COL_WARN}; border: none;")
            self.lbl_check_text.setText(f"SLOT {m['slot_char']} DIPILIH")
            self.lbl_desc_text.setText("Belum ada baseline kalibrasi. Pilih 'Kalibrasi Ulang' di menu Mesin Saya.")

    def _page_awam_detail_mesin(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(6, 6, 6, 6)
        lay.setSpacing(8)

        lay.addStretch(1)

        self.lbl_detail_icon = QLabel("⚙️")
        self.lbl_detail_icon.setAlignment(Qt.AlignCenter)
        self.lbl_detail_icon.setStyleSheet(f"font-size: 36px; color: {COL_ACCENT}; border: none;")
        lay.addWidget(self.lbl_detail_icon)

        self.lbl_detail_nama = QLabel("- Mesin -")
        self.lbl_detail_nama.setAlignment(Qt.AlignCenter)
        self.lbl_detail_nama.setWordWrap(True)
        self.lbl_detail_nama.setStyleSheet(f"font-size: 12px; font-weight: bold; color: {COL_TEXT_LIGHT}; border: none;")
        lay.addWidget(self.lbl_detail_nama)

        self.lbl_detail_info = QLabel("Slot [-]")
        self.lbl_detail_info.setAlignment(Qt.AlignCenter)
        self.lbl_detail_info.setStyleSheet(f"font-size: 9px; color: {COL_TEXT_DIM}; border: none;")
        lay.addWidget(self.lbl_detail_info)

        lay.addStretch(1)

        self.btn_detail_cek = QPushButton("🔍 Cek Sekarang")
        self.btn_detail_cek.setFixedHeight(42)
        self.btn_detail_cek.setStyleSheet(
            f"QPushButton {{ background-color: {COL_ACCENT}; color: #06131c; font-size: 11px; "
            f"font-weight: bold; border-radius: 6px; }}"
            f"QPushButton:hover {{ background-color: #5ecdfb; }}"
        )
        self.btn_detail_cek.clicked.connect(self._start_detail_cek)
        lay.addWidget(self.btn_detail_cek)

        self.btn_detail_kalib = QPushButton("🛠 Kalibrasi Ulang")
        self.btn_detail_kalib.setFixedHeight(42)
        self.btn_detail_kalib.clicked.connect(self._start_detail_kalib)
        lay.addWidget(self.btn_detail_kalib)

        self.btn_detail_hapus = QPushButton("🗑 Hapus Mesin Ini")
        self.btn_detail_hapus.setFixedHeight(30)
        self.btn_detail_hapus.setStyleSheet(
            f"QPushButton {{ background-color: {COL_PANEL_DARK}; color: {COL_BAD}; "
            f"border: 1px solid {COL_BAD}; border-radius: 4px; font-size: 9px; }}"
            f"QPushButton:hover {{ background-color: #2a1518; }}"
        )
        self.btn_detail_hapus.clicked.connect(self._delete_detail_machine)
        lay.addWidget(self.btn_detail_hapus)

        btn_kembali = QPushButton("← Kembali ke Mesin Saya")
        btn_kembali.setFixedHeight(26)
        btn_kembali.setStyleSheet(
            f"QPushButton {{ background-color: {COL_PANEL_DARK}; color: {COL_TEXT_DIM}; "
            f"border: 1px solid #2a3542; border-radius: 4px; font-size: 8px; }}"
            f"QPushButton:hover {{ background-color: {COL_ACCENT_DIM}; }}"
        )
        btn_kembali.clicked.connect(lambda: self._change_page(1))
        lay.addWidget(btn_kembali)

        self.detail_machine_idx = -1
        return page

    def _open_machine_detail(self, idx):
        if self._sesi_sedang_aktif():
            self._change_page(10)
            return
        self.detail_machine_idx = idx
        m = self.machines[idx]

        self.lbl_detail_icon.setText(m.get("icon", "⚙️"))
        self.lbl_detail_nama.setText(m["name"])

        terkunci = m.get("no_calib", False)
        mode_str = "Preset Pabrik (Terkunci)" if terkunci else "Dengan Kalibrasi"
        self.lbl_detail_info.setText(f"Slot [{m['slot_char']}] — {mode_str}")

        self.btn_detail_kalib.setEnabled(not terkunci)
        if terkunci:
            self.btn_detail_kalib.setText("🔒 Kalibrasi Terkunci (Preset Pabrik)")
            self.btn_detail_kalib.setStyleSheet(
                "QPushButton { background-color: #3a3f47; color: #7c8592; font-size: 10px; "
                "font-weight: bold; border-radius: 6px; }"
            )
        else:
            self.btn_detail_kalib.setText("🛠 Kalibrasi Ulang")
            self.btn_detail_kalib.setStyleSheet(
                f"QPushButton {{ background-color: {COL_WARN}; color: #1c1400; font-size: 11px; "
                f"font-weight: bold; border-radius: 6px; }}"
                f"QPushButton:hover {{ background-color: #ffd166; }}"
            )

        self.btn_detail_hapus.setVisible(idx >= 2)

        self._change_page(9)

    def _delete_detail_machine(self):
        idx = self.detail_machine_idx
        if idx < 2:
            return
        m = self.machines[idx]
        jawaban = QMessageBox.question(
            self, "Hapus Mesin",
            f"Yakin mau hapus '{m['name']}' dari daftar Mesin Saya?\n\n"
            f"Ini cuma menghapus catatannya di dashboard -- data hasil Cek "
            f"lama yang sudah ada di Riwayat TIDAK ikut terhapus.",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if jawaban != QMessageBox.Yes:
            return

        self.machines.pop(idx)
        self._save_machines_config()
        if self.selected_machine_idx == idx:
            self.selected_machine_idx = -1
            self.lbl_machine_active.setText("⚙️ - Pilih Slot -")
        elif self.selected_machine_idx > idx:
            self.selected_machine_idx -= 1

        if hasattr(self, 'awam_machine_grid'):
            self._rebuild_awam_machine_grid()
        if hasattr(self, 'machine_grid'):
            self._rebuild_machine_grid()
        self._refresh_wajah_alat()
        self._change_page(1)

    def _start_detail_cek(self):
        idx = self.detail_machine_idx
        if idx < 0:
            return
        self._select_machine_slot(idx, auto_calibrate=False)
        self._trigger_check_session()

    def _start_detail_kalib(self):
        idx = self.detail_machine_idx
        if idx < 0:
            return
        m = self.machines[idx]
        if m.get("no_calib", False):
            QMessageBox.information(self, "Terkunci", "Slot ini pakai preset pabrik yang terkunci -- tidak bisa dikalibrasi ulang lewat dashboard maupun firmware.")
            return
        self._select_machine_slot(idx, auto_calibrate=False)
        self._show_proses_loading("MEMPERSIAPKAN KALIBRASI", lambda: self._start_delayed_calibration(idx))

    def _page_awam_proses(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(6, 6, 6, 6)
        lay.setSpacing(10)

        lay.addStretch(1)

        self.lbl_proses_judul = QLabel("MEMPROSES...")
        self.lbl_proses_judul.setAlignment(Qt.AlignCenter)
        self.lbl_proses_judul.setStyleSheet(f"font-size: 11px; font-weight: bold; color: {COL_TEXT_LIGHT}; border: none;")
        lay.addWidget(self.lbl_proses_judul)

        self.lbl_proses_angka = QLabel("--")
        self.lbl_proses_angka.setAlignment(Qt.AlignCenter)
        self.lbl_proses_angka.setStyleSheet(f"font-size: 52px; font-weight: bold; color: {COL_ACCENT}; border: none;")
        lay.addWidget(self.lbl_proses_angka)

        self.lbl_proses_ket = QLabel("Mohon tunggu sebentar...")
        self.lbl_proses_ket.setAlignment(Qt.AlignCenter)
        self.lbl_proses_ket.setWordWrap(True)
        self.lbl_proses_ket.setStyleSheet(f"font-size: 9px; color: {COL_TEXT_DIM}; border: none;")
        lay.addWidget(self.lbl_proses_ket)

        lay.addStretch(1)

        self.btn_proses_batal = QPushButton("✖ Batalkan")
        self.btn_proses_batal.setFixedHeight(34)
        self.btn_proses_batal.setStyleSheet(
            f"QPushButton {{ background-color: {COL_PANEL_DARK}; color: {COL_BAD}; "
            f"border: 1px solid {COL_BAD}; border-radius: 6px; font-size: 10px; font-weight: bold; }}"
            f"QPushButton:hover {{ background-color: #2a1518; }}"
        )
        self.btn_proses_batal.clicked.connect(self._cancel_proses)
        lay.addWidget(self.btn_proses_batal)

        return page

    def _set_proses_labels(self, judul, angka, ket):
        if not hasattr(self, 'lbl_proses_judul'):
            return
        self.lbl_proses_judul.setText(judul)
        self.lbl_proses_angka.setText(angka)
        self.lbl_proses_ket.setText(ket)

    def _show_proses_loading(self, judul, on_selesai, durasi_detik=3):
        if not self.is_technician_mode:
            self._change_page(10)
        self._loading_dots_step = 0
        self._loading_dots_total_tick = durasi_detik * 2  
        self._loading_on_selesai = on_selesai
        self._set_proses_labels(judul, "", "Menyiapkan...")

        if hasattr(self, 'loading_dots_timer') and self.loading_dots_timer.isActive():
            self.loading_dots_timer.stop()
        self.loading_dots_timer = QTimer(self)
        self.loading_dots_timer.timeout.connect(self._loading_dots_tick)
        self.loading_dots_timer.start(500)

    def _loading_dots_tick(self):
        self._loading_dots_step += 1
        dots = "." * ((self._loading_dots_step % 3) + 1)
        self.lbl_proses_angka.setText(dots)
        if self._loading_dots_step >= self._loading_dots_total_tick:
            self.loading_dots_timer.stop()
            callback = self._loading_on_selesai
            self._loading_on_selesai = None
            if callback:
                callback()

    def _cancel_proses(self):
        if hasattr(self, 'loading_dots_timer') and self.loading_dots_timer.isActive():
            self.loading_dots_timer.stop()
        self._loading_on_selesai = None
        if hasattr(self, 'delay_countdown_timer') and self.delay_countdown_timer.isActive():
            self.delay_countdown_timer.stop()
        if self.delay_calibration_timer.isActive():
            self.delay_calibration_timer.stop()
        if self.calibration_timer.isActive():
            self.calibration_timer.stop()
        if self.check_session_timer.isActive():
            self.check_session_timer.stop()

        self.pending_calibration_idx = -1
        self.calibrating_machine_idx = -1
        self.is_check_session_active = False

        if self.detail_machine_idx >= 0:
            self._change_page(9)
        else:
            self._change_page(1)

    def _start_delayed_calibration(self, idx):
        if self.calibration_timer.isActive():
            self.calibration_timer.stop()
        if self.delay_calibration_timer.isActive():
            self.delay_calibration_timer.stop()

        self.pending_calibration_idx = idx
        self.delay_seconds_left = 4

        self._set_proses_labels("PERSIAPAN KALIBRASI", str(self.delay_seconds_left), "Menunggu kestabilan perangkat keras di slot aktif...")
        if not self.is_technician_mode:
            self._change_page(10)

        self.delay_countdown_timer = QTimer(self)
        self.delay_countdown_timer.timeout.connect(self._delay_tick)
        self.delay_countdown_timer.start(1000)

        self.delay_calibration_timer.start(4000)

    def _delay_tick(self):
        self.delay_seconds_left -= 1
        if self.delay_seconds_left > 0:
            self._set_proses_labels("PERSIAPAN KALIBRASI", str(self.delay_seconds_left), "Menunggu kestabilan perangkat keras di slot aktif...")
        else:
            if hasattr(self, 'delay_countdown_timer'):
                self.delay_countdown_timer.stop()

    def _execute_actual_calibration(self):
        if hasattr(self, 'delay_countdown_timer'):
            self.delay_countdown_timer.stop()

        idx = self.pending_calibration_idx
        if idx < 0:
            return

        self.calibrating_machine_idx = idx
        self.calibration_time_left = 180
        self.calibration_timer.start(1000)

        self._send_command('R')
        time.sleep(0.05)
        self._send_command(self.machines[idx]["slot_char"])

        mins = self.calibration_time_left // 60
        secs = self.calibration_time_left % 60
        self._set_proses_labels("KALIBRASI BERJALAN", f"{mins:02d}:{secs:02d}", "Jaga mesin tetap berjalan normal selama proses ini.")

    def _calibration_countdown_tick(self):
        if self.calibration_time_left > 0:
            self.calibration_time_left -= 1
            mins = self.calibration_time_left // 60
            secs = self.calibration_time_left % 60
            self._set_proses_labels("KALIBRASI BERJALAN", f"{mins:02d}:{secs:02d}", "Jaga mesin tetap berjalan normal selama proses ini.")
        else:
            self.calibration_timer.stop()
            self.calibrating_machine_idx = -1
            self.lbl_check_text.setStyleSheet(f"font-size: 10px; font-weight: bold; color: {COL_OK}; border: none;")
            self.lbl_check_text.setText("KALIBRASI SELESAI")
            self.lbl_desc_text.setText("Baseline D² tersimpan di slot aktif.")
            if not self.is_technician_mode:
                QMessageBox.information(self, "Kalibrasi Selesai", "Kalibrasi selesai. Baseline baru sudah tersimpan di slot ini.")
                self._change_page(9 if self.detail_machine_idx >= 0 else 1)

    def _page_awam_cek_status(self):
        page = QWidget()
        outer = QVBoxLayout(page)
        outer.setContentsMargins(4, 4, 4, 4)
        outer.setSpacing(6)

        split = QHBoxLayout()
        split.setSpacing(6)

        box_left = QFrame()
        box_left.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px solid #2a3542; border-radius: 6px;")
        left_lay = QVBoxLayout(box_left)
        left_lay.setAlignment(Qt.AlignCenter)
        left_lay.setContentsMargins(4, 8, 4, 8)

        self.lbl_beranda_icon = QLabel("❔")
        self.lbl_beranda_icon.setAlignment(Qt.AlignCenter)
        self.lbl_beranda_icon.setStyleSheet("font-size: 46px; color: #a0aec0; border: none;")
        left_lay.addWidget(self.lbl_beranda_icon)

        self.lbl_check_text = QLabel("BELUM ADA HASIL CHECK")
        self.lbl_check_text.setAlignment(Qt.AlignCenter)
        self.lbl_check_text.setWordWrap(True)
        self.lbl_check_text.setStyleSheet("font-size: 9px; font-weight: bold; color: #e2e8f0; border: none;")
        left_lay.addWidget(self.lbl_check_text)

        split.addWidget(box_left, 1)

        box_right = QFrame()
        box_right.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px solid #2a3542; border-radius: 6px;")
        right_lay = QVBoxLayout(box_right)
        right_lay.setContentsMargins(10, 8, 10, 8)
        right_lay.setSpacing(6)

        lbl_rincian_title = QLabel("RINCIAN HASIL CEK")
        lbl_rincian_title.setStyleSheet(f"font-size: 8px; font-weight: bold; color: {COL_TEXT_DIM}; border: none;")
        right_lay.addWidget(lbl_rincian_title)

        def _baris_rincian(judul, attr_name):
            row = QHBoxLayout()
            lbl_j = QLabel(judul)
            lbl_j.setStyleSheet(f"font-size: 9px; color: {COL_TEXT_DIM}; border: none;")
            lbl_v = QLabel("--")
            lbl_v.setAlignment(Qt.AlignRight)
            lbl_v.setStyleSheet(f"font-size: 11px; font-weight: bold; color: {COL_TEXT_LIGHT}; border: none;")
            row.addWidget(lbl_j, 1)
            row.addWidget(lbl_v, 1)
            setattr(self, attr_name, lbl_v)
            right_lay.addLayout(row)

        _baris_rincian("📳 Getaran", "lbl_rincian_getaran")
        _baris_rincian("🔊 Suara", "lbl_rincian_suara")
        _baris_rincian("🌡️ Suhu", "lbl_rincian_suhu")
        _baris_rincian("🕐 Waktu Cek", "lbl_rincian_waktu")
        _baris_rincian("🔧 Jenis Masalah", "lbl_rincian_diagnosis")

        right_lay.addStretch(1)
        split.addWidget(box_right, 2)

        outer.addLayout(split, 3)

        box_desc = QFrame()
        box_desc.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px solid #2a3542; border-radius: 6px;")
        d_lay = QVBoxLayout(box_desc)
        d_lay.setAlignment(Qt.AlignCenter)
        d_lay.setContentsMargins(4, 4, 4, 4)

        self.lbl_desc_text = QLabel("Pilih mesin di menu MESIN SAYA, lalu tekan Cek Sekarang.")
        self.lbl_desc_text.setAlignment(Qt.AlignCenter)
        self.lbl_desc_text.setWordWrap(True)
        self.lbl_desc_text.setStyleSheet(f"font-size: 9px; color: {COL_TEXT_DIM}; border: none;")
        d_lay.addWidget(self.lbl_desc_text)
        outer.addWidget(box_desc, 1)

        self.lbl_estimasi_servis = QLabel("Estimasi servis berikutnya: --")
        self.lbl_estimasi_servis.setAlignment(Qt.AlignCenter)
        self.lbl_estimasi_servis.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {COL_TEXT_LIGHT}; background-color: {COL_PANEL_DARK}; padding: 4px; border-radius: 4px; border: 1px solid #2a3542;")
        outer.addWidget(self.lbl_estimasi_servis)

        return page

    def _page_raw(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(2)

        main_raw = QHBoxLayout()
        main_raw.setSpacing(2)
        
        layout_grafik_grid = QGridLayout()
        layout_grafik_grid.setSpacing(2)
        layout_grafik_grid.setContentsMargins(0, 0, 0, 0)
        
        pg.setConfigOptions(antialias=True)

        def _tidy_plot(plot_widget, title):
            axis_font = QFont("Arial", 7)
            plot_item = plot_widget.getPlotItem()
            for axis_name in ("left", "bottom"):
                axis = plot_item.getAxis(axis_name)
                axis.setStyle(tickFont=axis_font, tickTextOffset=2, autoExpandTextSpace=True)
                axis.setTextPen(pg.mkPen('#dddddd'))
            plot_item.getAxis("bottom").setTickSpacing(major=15, minor=15)
            plot_item.getAxis("left").setWidth(30)
            plot_widget.setTitle(title, size="8pt")
            plot_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.graph_a = pg.PlotWidget(title="Sound (dB)")
        self.graph_a.setBackground(COL_PANEL_DARK)
        self.graph_a.showGrid(x=True, y=True, alpha=0.2)
        self.curve_a = self.graph_a.plot(pen=pg.mkPen('#a78bfa', width=1.2))
        _tidy_plot(self.graph_a, "Sound (dB)")
        layout_grafik_grid.addWidget(self.graph_a, 0, 0)
        
        self.graph_temp = pg.PlotWidget(title="Temp (°C)")
        self.graph_temp.setBackground(COL_PANEL_DARK)
        self.graph_temp.showGrid(x=True, y=True, alpha=0.2)
        self.curve_temp = self.graph_temp.plot(pen=pg.mkPen('#f472b6', width=1.2))
        _tidy_plot(self.graph_temp, "Temp (°C)")
        layout_grafik_grid.addWidget(self.graph_temp, 0, 1)

        self.graph_v = pg.PlotWidget(title="Vibration (G)")
        self.graph_v.setBackground(COL_PANEL_DARK)
        self.graph_v.showGrid(x=True, y=True, alpha=0.2)
        self.curve_v = self.graph_v.plot(pen=pg.mkPen('#818cf8', width=1.2))
        _tidy_plot(self.graph_v, "Vibration (G)")
        layout_grafik_grid.addWidget(self.graph_v, 1, 0)
        
        self.panel_ai = QFrame()
        self.panel_ai.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border-radius: 3px; border: 1px solid #2a3542;")
        lay_ai = QVBoxLayout(self.panel_ai)
        lay_ai.setContentsMargins(4, 2, 4, 2)
        lay_ai.setSpacing(1)

        lbl_ai_title = QLabel("🤖 AI DIAGNOSTICS & GROUND TRUTH")
        lbl_ai_title.setStyleSheet(f"font-size: 7px; font-weight: bold; color: {COL_ACCENT};")
        lay_ai.addWidget(lbl_ai_title)

        grid_ai = QGridLayout()
        grid_ai.setHorizontalSpacing(2)
        grid_ai.setVerticalSpacing(0)

        def _ai_row(r, label_text):
            lbl1 = QLabel(label_text)
            lbl1.setStyleSheet("font-size: 7px; color: #999;")
            lbl2 = QLabel("--")
            lbl2.setStyleSheet("font-size: 8px; font-weight: bold; color: #eee;")
            lbl2.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            grid_ai.addWidget(lbl1, r, 0)
            grid_ai.addWidget(lbl2, r, 1)
            return lbl2

        self.lbl_hs_val = _ai_row(0, "Health Score:")
        self.lbl_ml_val = _ai_row(1, "TinyML Mode:")
        self.lbl_diag_val = _ai_row(2, "Fault Diag:")
        self.lbl_kurt_val = _ai_row(3, "Kurtosis:")
        self.lbl_flags_val = _ai_row(4, "Multi-Fault:")

        self.prog_hs = QProgressBar()
        self.prog_hs.setFixedHeight(3)
        self.prog_hs.setTextVisible(False)
        self.prog_hs.setStyleSheet(
            "QProgressBar { border: none; background-color: #333; border-radius: 1px; }"
            "QProgressBar::chunk { background-color: " + COL_OK + "; border-radius: 1px; }"
        )
        grid_ai.addWidget(self.prog_hs, 5, 0, 1, 2)

        gt_layout = QHBoxLayout()
        gt_layout.setSpacing(1)
        for gt_char, gt_label in [('O', 'NORM'), ('U', 'UNB'), ('M', 'MIS'), ('F', 'BRG'), ('L', 'LUB'), ('D', 'MATI')]:
            b = QPushButton(gt_label)
            b.setStyleSheet("font-size: 6px; background-color: #1c2027; color: #bbb; border: 1px solid #444; border-radius: 2px;")
            b.clicked.connect(lambda checked, ch=gt_char: self._send_command(ch))
            gt_layout.addWidget(b)
        grid_ai.addLayout(gt_layout, 6, 0, 1, 2)

        lay_ai.addLayout(grid_ai)
        layout_grafik_grid.addWidget(self.panel_ai, 1, 1)

        layout_grafik_grid.setColumnStretch(0, 1)
        layout_grafik_grid.setColumnStretch(1, 1)
        layout_grafik_grid.setRowStretch(0, 1)
        layout_grafik_grid.setRowStretch(1, 1)
        
        main_raw.addLayout(layout_grafik_grid, 6)

        panel_kanan = QVBoxLayout()
        panel_kanan.setSpacing(1)

        frame_status = QFrame()
        frame_status.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border-radius: 3px; border: 1px solid #2a3542;")
        fs_lay = QVBoxLayout(frame_status)
        fs_lay.setContentsMargins(4, 2, 4, 2)
        fs_lay.setSpacing(1)

        self.lbl_sys_status = QLabel("● STANDBY")
        self.lbl_sys_status.setStyleSheet("font-size: 8px; font-weight: bold; color: #888888; padding-bottom: 2px;")
        fs_lay.addWidget(self.lbl_sys_status)

        row_style = "border-bottom: 1px solid #33363b;"
        name_style = "font-size: 7px; color: #999999;"
        val_style = "font-size: 9px; color: #eeeeee; font-weight: bold;"

        grid_val = QGridLayout()
        grid_val.setContentsMargins(0, 0, 0, 0)
        grid_val.setHorizontalSpacing(2)
        grid_val.setVerticalSpacing(1)
        grid_val.setColumnStretch(0, 1)
        grid_val.setColumnStretch(1, 1)

        def _param_row(row, name):
            lbl_name = QLabel(name)
            lbl_name.setStyleSheet(name_style + row_style + "padding: 1px 0px;")
            lbl_val = QLabel("--")
            lbl_val.setStyleSheet(val_style + row_style + "padding: 1px 0px;")
            lbl_val.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            grid_val.addWidget(lbl_name, row, 0)
            grid_val.addWidget(lbl_val, row, 1)
            return lbl_val

        self.lbl_val_v = _param_row(0, "Vibration")
        self.lbl_val_a = _param_row(1, "Sound")
        self.lbl_val_temp = _param_row(2, "Temp")
        self.lbl_val_rpm = _param_row(3, "RPM")
        self.lbl_val_d2 = _param_row(4, "D² (Severity)")

        fs_lay.addLayout(grid_val)

        self.lbl_val_vxyz = QLabel("(X: -- | Y: -- | Z: -- G)")
        self.lbl_val_vxyz.setStyleSheet("font-size: 7px; color: #777777; padding: 2px 0px 0px 0px;")
        self.lbl_val_vxyz.setWordWrap(True)
        fs_lay.addWidget(self.lbl_val_vxyz)

        panel_kanan.addWidget(frame_status)
        main_raw.addLayout(panel_kanan, 3)

        layout.addLayout(main_raw)
        return page

    def _page_recording(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(2)

        rec_kondisi_lay = QHBoxLayout()
        lbl_kondisi = QLabel("Kondisi:")
        lbl_kondisi.setStyleSheet("font-size: 8px; color: #aaa;")
        self.cmb_rec_kondisi = QComboBox()
        for _huruf, _label in KONDISI_REKAM_OPSI:
            self.cmb_rec_kondisi.addItem(_label)
        self.cmb_rec_kondisi.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: {COL_TEXT_LIGHT}; font-size: 8px;")
        rec_kondisi_lay.addWidget(lbl_kondisi)
        rec_kondisi_lay.addWidget(self.cmb_rec_kondisi, 1)
        layout.addLayout(rec_kondisi_lay)

        rec_dataset_ctrl = QHBoxLayout()
        self.btn_toggle_dataset_rec = QPushButton("💾 SIMPAN KE DATASET (Format Logger)")
        self.btn_toggle_dataset_rec.setStyleSheet("background-color: #2f7d4f; color: #ffffff; font-weight: bold; font-size: 9px; height: 24px; border-radius: 3px;")
        self.btn_toggle_dataset_rec.clicked.connect(self._toggle_dataset_recording)

        self.lbl_dataset_rec_status = QLabel("● Belum Menulis")
        self.lbl_dataset_rec_status.setStyleSheet("font-size: 8px; color: #aaa;")

        rec_dataset_ctrl.addWidget(self.btn_toggle_dataset_rec, 2)
        rec_dataset_ctrl.addWidget(self.lbl_dataset_rec_status, 1)
        layout.addLayout(rec_dataset_ctrl)

        lbl_dataset_hint = QLabel("* File tersimpan di folder Dataset/, format & nama file sama seperti loggerserial.py.")
        lbl_dataset_hint.setStyleSheet("font-size: 7px; color: #888; font-style: italic;")
        lbl_dataset_hint.setWordWrap(True)
        layout.addWidget(lbl_dataset_hint)

        rec_ctrl = QHBoxLayout()
        self.btn_toggle_rec = QPushButton("MULAI RECORDING")
        self.btn_toggle_rec.setStyleSheet("background-color: #cfcfcf; color: #000000; font-weight: bold; font-size: 9px; height: 24px; border-radius: 3px;")
        self.btn_toggle_rec.clicked.connect(self._toggle_recording)

        self.lbl_rec_status = QLabel("● Device Standby")
        self.lbl_rec_status.setStyleSheet("font-size: 8px; color: #aaa;")

        rec_ctrl.addWidget(self.btn_toggle_rec, 2)
        rec_ctrl.addWidget(self.lbl_rec_status, 1)
        layout.addLayout(rec_ctrl)

        lbl_daftar = QLabel("Daftar Rekaman Mesin (.CSV):")
        lbl_daftar.setStyleSheet("font-size: 8px;")
        layout.addWidget(lbl_daftar)
        
        self.log_list = QListWidget()
        self.log_list.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: white; font-size: 9px;")
        self.log_list.itemDoubleClicked.connect(self._open_log_detail)
        layout.addWidget(self.log_list)

        btn_lay = QHBoxLayout()
        self.btn_watch_log = QPushButton("Buka Panel Deteksi")
        self.btn_export_excel = QPushButton("Export ke Excel")
        self.btn_delete_log = QPushButton("Hapus Rekaman")

        self.btn_watch_log.setStyleSheet("background-color: #cfcfcf; color: #000000; font-size: 9px; height: 22px; font-weight: bold; border-radius: 2px;")
        self.btn_export_excel.setStyleSheet("background-color: #217346; color: #ffffff; font-size: 9px; height: 22px; font-weight: bold; border-radius: 2px;")
        self.btn_delete_log.setStyleSheet("background-color: #cfcfcf; color: #000000; font-size: 9px; height: 22px; font-weight: bold; border-radius: 2px;")

        self.btn_watch_log.clicked.connect(self._open_log_detail_from_button)
        self.btn_export_excel.clicked.connect(self._export_selected_log_to_excel)
        self.btn_delete_log.clicked.connect(self._delete_selected_log)

        btn_lay.addWidget(self.btn_watch_log)
        btn_lay.addWidget(self.btn_export_excel)
        btn_lay.addWidget(self.btn_delete_log)
        layout.addLayout(btn_lay)

        lbl_hint = QLabel("* Klik file rekaman untuk membuka panel forensik anomali.")
        lbl_hint.setStyleSheet("font-size: 8px; color: #888; font-style: italic;")
        layout.addWidget(lbl_hint)

        return page

    def _page_processed(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(2)

        self.lbl_proc_snapshot = QLabel("Menunggu data... | RPM: -- | D²: --")
        self.lbl_proc_snapshot.setStyleSheet("font-size: 8px; color: #999;")
        layout.addWidget(self.lbl_proc_snapshot)

        grid = QGridLayout()
        grid.setSpacing(2)
        pg.setConfigOptions(antialias=True)

        self.graph_fft = pg.PlotWidget(title="Frequency Spectrum (FFT)")
        self.graph_fft.setBackground(COL_PANEL_DARK)
        self.graph_fft.showGrid(x=True, y=True, alpha=0.3)
        self.curve_fft = self.graph_fft.plot(pen=pg.mkPen('#a855f7', width=1.2))
        grid.addWidget(self.graph_fft, 0, 0, 1, 2)

        self.graph_rpm = pg.PlotWidget(title="RPM Estimasi")
        self.graph_rpm.setBackground(COL_PANEL_DARK)
        self.graph_rpm.showGrid(x=True, y=True, alpha=0.3)
        self.curve_rpm = self.graph_rpm.plot(pen=pg.mkPen('#4da6ff', width=1.2))
        grid.addWidget(self.graph_rpm, 1, 0)

        self.graph_d2 = pg.PlotWidget(title="Mahalanobis D²")
        self.graph_d2.setBackground(COL_PANEL_DARK)
        self.graph_d2.showGrid(x=True, y=True, alpha=0.3)
        self.curve_d2 = self.graph_d2.plot(pen=pg.mkPen('#ff6666', width=1.2))
        
        line_waspada = pg.InfiniteLine(pos=D2_THRESHOLD_WASPADA, angle=0, pen=pg.mkPen(COL_WARN, width=1, style=Qt.DashLine))
        line_bahaya = pg.InfiniteLine(pos=D2_THRESHOLD_BAHAYA, angle=0, pen=pg.mkPen(COL_BAD, width=1, style=Qt.DashLine))
        self.graph_d2.addItem(line_waspada)
        self.graph_d2.addItem(line_bahaya)
        grid.addWidget(self.graph_d2, 1, 1)

        layout.addLayout(grid, 3)

        lbl_anomali_title = QLabel("Log Kejadian Anomali:")
        lbl_anomali_title.setStyleSheet("font-size: 8px; font-weight: bold; color: #ccc;")
        layout.addWidget(lbl_anomali_title)

        self.list_anomali = QListWidget()
        self.list_anomali.setStyleSheet("background-color: #ffffff; color: #222222; font-size: 9px;")
        self.list_anomali.addItem("Tidak ada kejadian anomali sepanjang sesi ini.")
        layout.addWidget(self.list_anomali, 2)

        self.lbl_session_summary = QLabel(
            "Sesi: 0 sample | RPM rata-rata: 0.0 | D² max: 0.00 | Kondisi terparah: Normal | Waspada: 0x, Bahaya: 0x."
        )
        self.lbl_session_summary.setStyleSheet(
            "font-size: 9px; color: #1c3d1c; background-color: #d7f0d7; padding: 3px; border-radius: 2px;"
        )
        self.lbl_session_summary.setWordWrap(True)
        layout.addWidget(self.lbl_session_summary)

        return page

    def _page_summary(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(4)

        self.lbl_sum_machine = QLabel("Target: Belum dipilih")
        self.lbl_sum_machine.setStyleSheet(f"font-size: 11px; font-weight: bold; color: {COL_TEXT_LIGHT};")
        layout.addWidget(self.lbl_sum_machine)

        grid = QGridLayout()
        grid.setSpacing(4)

        self.gauge_s = GradientGauge("Sound", 30.0, 100.0, 75.0, 85.0, "dB")
        self.gauge_t = GradientGauge("Temperature", 20.0, 80.0, 42.0, 50.0, "°C")
        self.gauge_v = GradientGauge("Vibration", 0.0, 0.5, 0.18, 0.25, "G")

        grid.addWidget(self.gauge_s, 0, 0)
        grid.addWidget(self.gauge_t, 0, 1)
        grid.addWidget(self.gauge_v, 1, 0)

        self.reserved_slot_sum = QFrame()
        self.reserved_slot_sum.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px dashed {COL_TEXT_DIM}; border-radius: 3px;")
        grid.addWidget(self.reserved_slot_sum, 1, 1)

        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)
        grid.setRowStretch(0, 1)
        grid.setRowStretch(1, 1)

        layout.addLayout(grid)
        layout.addStretch(1)

        self.lbl_diag_desc_summary = QLabel("Menunggu data deteksi...")
        self.lbl_diag_desc_summary.setStyleSheet(f"font-size: 9px; color: {COL_TEXT_DIM}; border-top: 1px solid #333; padding-top: 3px;")
        self.lbl_diag_desc_summary.setWordWrap(True)
        self.lbl_diag_desc_summary.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.lbl_diag_desc_summary)

        return page

    def _page_machine_select(self):
        page = QWidget()
        self.machine_page_layout = QVBoxLayout(page)
        self.machine_page_layout.setContentsMargins(4, 4, 4, 4)
        self.machine_page_layout.setSpacing(2)

        top_row = QHBoxLayout()
        lbl_title = QLabel("Manajemen Slot Mesin (0-9)")
        lbl_title.setStyleSheet(f"font-size: 10px; font-weight: bold; color: {COL_TEXT_LIGHT};")
        top_row.addWidget(lbl_title)
        top_row.addStretch(1)

        self.btn_add_machine_modal = QPushButton("➕ TAMBAH SLOT")
        self.btn_add_machine_modal.setStyleSheet(f"background-color: {COL_ACCENT}; color: #101216; font-weight: bold; font-size: 8px; padding: 3px; border-radius: 2px;")
        self.btn_add_machine_modal.clicked.connect(self._open_add_machine_dialog)
        top_row.addWidget(self.btn_add_machine_modal)

        self.btn_machine_delete_mode = QPushButton("🗑 HAPUS")
        self.btn_machine_delete_mode.setCheckable(True)
        self.btn_machine_delete_mode.setStyleSheet(self._delete_mode_btn_style(False))
        self.btn_machine_delete_mode.clicked.connect(self._toggle_machine_delete_mode)
        top_row.addWidget(self.btn_machine_delete_mode)
        
        self.machine_page_layout.addLayout(top_row)

        self.machine_grid = QGridLayout()
        self.machine_grid.setSpacing(3)
        for c in range(3):
            self.machine_grid.setColumnStretch(c, 1)
        for r in range(2):
            self.machine_grid.setRowStretch(r, 1)
        self.machine_page_layout.addLayout(self.machine_grid, 1)

        self._rebuild_machine_grid()
        return page

    def _delete_mode_btn_style(self, active):
        if active:
            return f"background-color: {COL_BAD}; color: #101216; font-weight: bold; font-size: 8px; padding: 3px; border-radius: 2px;"
        return f"background-color: transparent; color: {COL_BAD}; font-weight: bold; font-size: 8px; padding: 3px; border: 1px solid {COL_BAD}; border-radius: 2px;"

    def _toggle_machine_delete_mode(self):
        self.machine_delete_mode = self.btn_machine_delete_mode.isChecked()
        self.btn_machine_delete_mode.setStyleSheet(self._delete_mode_btn_style(self.machine_delete_mode))

    def _make_machine_card(self, idx):
        btn = QPushButton()
        btn.setFixedHeight(55)
        lay = QVBoxLayout(btn)
        lay.setContentsMargins(1, 1, 1, 1)
        lay.setSpacing(0)

        m = self.machines[idx]
        lbl_icon = QLabel(m["icon"])
        lbl_name = QLabel(m["name"])
        mode_str = "Tanpa Kalibrasi" if m.get("no_calib", False) else f"Slot [{m['slot_char']}]"
        lbl_cluster = QLabel(f"⚡ {mode_str}")
        
        btn.clicked.connect(lambda checked, i=idx: self._on_machine_card_clicked(i))
        selected = (idx == self.selected_machine_idx)
        border_col = COL_ACCENT if selected else "#2a3542"
        bg = COL_ACCENT_DIM if selected else COL_PANEL_DARK

        lbl_icon.setAlignment(Qt.AlignCenter)
        lbl_icon.setStyleSheet("font-size: 14px; background: transparent; border: none;")
        
        lbl_name.setAlignment(Qt.AlignCenter)
        lbl_name.setWordWrap(True)
        lbl_name.setStyleSheet(f"font-size: 8px; font-weight: bold; color: {COL_TEXT_LIGHT}; background: transparent; border: none;")

        lbl_cluster.setAlignment(Qt.AlignCenter)
        cluster_color = COL_OK if idx < 2 else COL_ACCENT
        lbl_cluster.setStyleSheet(f"font-size: 7px; color: {cluster_color}; background: transparent; border: none;")

        lay.addWidget(lbl_icon)
        lay.addWidget(lbl_name)
        lay.addWidget(lbl_cluster)

        btn.setStyleSheet(
            f"QPushButton {{ background-color: {bg}; border: 1px solid {border_col}; border-radius: 3px; }}"
            f"QPushButton:hover {{ border: 1px solid {COL_ACCENT}; }}"
        )
        return btn

    def _rebuild_machine_grid(self):
        while self.machine_grid.count():
            item = self.machine_grid.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

        for i in range(len(self.machines)):
            self.machine_grid.addWidget(self._make_machine_card(i), i // 3, i % 3)

    def _open_add_machine_dialog(self):
        dlg = AddMachineDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            new_data = dlg.get_data()
            new_idx = len(self.machines)
            self.machines.append(new_data)
            self._save_machines_config()  # Simpan ke memori lokal
            self._rebuild_machine_grid()
            if hasattr(self, 'awam_machine_grid'):
                self._rebuild_awam_machine_grid()
            self._select_machine_slot(new_idx)

    def _on_machine_card_clicked(self, idx):
        if self.machine_delete_mode:
            if idx < 2:
                QMessageBox.warning(self, "Peringatan", "Slot utama (0 & 1) tidak dapat dihapus!")
                return
            self.machines.pop(idx)
            self._save_machines_config()  # Simpan ke memori lokal
            if self.selected_machine_idx == idx:
                self.selected_machine_idx = -1
                self.lbl_machine_active.setText("⚙️ - Pilih Slot -")
            elif self.selected_machine_idx > idx:
                self.selected_machine_idx -= 1
            self._rebuild_machine_grid()
            if hasattr(self, 'awam_machine_grid'):
                self._rebuild_awam_machine_grid()
            return

        self._select_machine_slot(idx)
        self._rebuild_machine_grid()

    def _page_log_detail(self):
        page = QWidget()
        root = QVBoxLayout(page)
        root.setContentsMargins(2, 2, 2, 2)
        root.setSpacing(2)

        self.lbl_logdet_title = QLabel("HASIL DETEKSI REKAMAN: -")
        self.lbl_logdet_title.setStyleSheet("font-size: 9px; font-weight: bold; color: #ffffff;")
        self.lbl_logdet_title.setWordWrap(True)
        root.addWidget(self.lbl_logdet_title)

        status_row = QHBoxLayout()
        status_row.setSpacing(2)

        self.lbl_logdet_dot = QLabel("●")
        self.lbl_logdet_dot.setStyleSheet("font-size: 12px; color: #888888;")
        status_row.addWidget(self.lbl_logdet_dot)

        self.lbl_logdet_peak = QLabel("Nilai puncak: -")
        self.lbl_logdet_peak.setWordWrap(True)
        self.lbl_logdet_peak.setStyleSheet("font-size: 8px; color: #cccccc;")
        status_row.addWidget(self.lbl_logdet_peak, 1)

        root.addLayout(status_row)

        grid = QGridLayout()
        grid.setSpacing(1)
        pg.setConfigOptions(antialias=True)

        self.graph_logdet_a = pg.PlotWidget(title="Sound (dB)")
        self.graph_logdet_a.setBackground(COL_PANEL_DARK)
        self.graph_logdet_a.showGrid(x=True, y=True, alpha=0.2)
        self.curve_logdet_a = self.graph_logdet_a.plot(pen=pg.mkPen('#4da6ff', width=1.2))
        grid.addWidget(self.graph_logdet_a, 0, 0)

        self.graph_logdet_temp = pg.PlotWidget(title="Temp (°C)")
        self.graph_logdet_temp.setBackground(COL_PANEL_DARK)
        self.graph_logdet_temp.showGrid(x=True, y=True, alpha=0.2)
        self.curve_logdet_temp = self.graph_logdet_temp.plot(pen=pg.mkPen('#e040fb', width=1.2))
        grid.addWidget(self.graph_logdet_temp, 0, 1)

        self.graph_logdet_v = pg.PlotWidget(title="Vibration (G)")
        self.graph_logdet_v.setBackground(COL_PANEL_DARK)
        self.graph_logdet_v.showGrid(x=True, y=True, alpha=0.2)
        self.curve_logdet_v = self.graph_logdet_v.plot(pen=pg.mkPen('#ff4d4d', width=1.2))
        grid.addWidget(self.graph_logdet_v, 1, 0)

        self.reserved_slot_logdet = QFrame()
        self.reserved_slot_logdet.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border: 1px dashed {COL_TEXT_DIM}; border-radius: 3px;")
        grid.addWidget(self.reserved_slot_logdet, 1, 1)

        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)

        root.addLayout(grid, 20)

        ctrl = QHBoxLayout()
        ctrl.setSpacing(2)

        self.btn_logdet_play_pause = QPushButton("▶ PLAY")
        self.speed_logdet_combo = QComboBox()
        self.speed_logdet_combo.addItems(["0.5x", "1x", "2x", "4x"])
        self.speed_logdet_combo.setCurrentIndex(1)
        self.btn_logdet_back = QPushButton("◄ KEMBALI")

        self.btn_logdet_play_pause.setStyleSheet("background-color: #cfcfcf; color: #000000; font-weight: bold; font-size: 9px; height: 24px; border-radius: 3px;")
        self.btn_logdet_back.setStyleSheet(f"background-color: {COL_BAD}; color: #ffffff; font-weight: bold; font-size: 9px; height: 24px; border-radius: 3px;")
        self.speed_logdet_combo.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: white; font-size: 9px;")

        self.btn_logdet_play_pause.clicked.connect(self._logdet_toggle_play)
        self.speed_logdet_combo.currentIndexChanged.connect(self._logdet_change_speed)
        self.btn_logdet_back.clicked.connect(self._logdet_back)

        self.slider_logdet = QSlider(Qt.Horizontal)
        self.slider_logdet.setMinimum(0)
        self.slider_logdet.setMaximum(0)
        self.slider_logdet.sliderMoved.connect(self._logdet_seek)

        self.lbl_logdet_pos = QLabel("0 / 0")
        self.lbl_logdet_pos.setStyleSheet("font-size: 8px; color: #aaa;")

        ctrl.addWidget(self.btn_logdet_play_pause, 2)
        ctrl.addWidget(self.speed_logdet_combo, 1)
        ctrl.addWidget(self.slider_logdet, 5)
        ctrl.addWidget(self.lbl_logdet_pos, 1)
        ctrl.addWidget(self.btn_logdet_back, 2)
        root.addLayout(ctrl)

        return page

    def _logdet_back(self):
        self.logdet_timer.stop()
        self._change_page(5)

    def _logdet_load_csv(self, filepath):
        times, v_vals, a_vals, t_vals = [], [], [], []
        try:
            with open(filepath, 'r') as f:
                reader = csv.reader(f)
                header = next(reader)
                idx_v = header.index('rms_v') if 'rms_v' in header else 2
                idx_a = header.index('rms_a') if 'rms_a' in header else 3
                idx_t = header.index('temp') if 'temp' in header else 4

                for i, row in enumerate(reader):
                    if len(row) <= max(idx_v, idx_a, idx_t): continue
                    times.append(i)
                    v_vals.append(float(row[idx_v]))
                    a_vals.append(float(row[idx_a]))
                    t_vals.append(float(row[idx_t]))
        except Exception as e:
            print(f"Error load csv: {e}")
        return times, v_vals, a_vals, t_vals

    def _logdet_render_diagnosis_summary(self):
        if not self.logdet_v_vals:
            self.lbl_logdet_dot.setStyleSheet("font-size: 12px; color: #888888;")
            self.lbl_logdet_peak.setText("Data kosong.")
            return

        peak_v = max(self.logdet_v_vals)
        peak_a = max(self.logdet_a_vals)
        peak_temp = max(self.logdet_t_vals)

        if peak_v > 0.25 or peak_temp > 50.0:
            status, col = "BAHAYA", COL_BAD
        elif peak_v > 0.18 or peak_temp > 42.0:
            status, col = "WASPADA", COL_WARN
        else:
            status, col = "NORMAL", COL_OK

        self.lbl_logdet_dot.setStyleSheet(f"font-size: 12px; color: {col};")
        self.lbl_logdet_peak.setText(
            f"{status} — Puncak: Vib {peak_v:.2f} G | Snd {peak_a:.1f} dB | Tmp {peak_temp:.1f} °C"
        )

    def _logdet_render_frame(self, idx):
        start = max(0, idx - 49)
        t = self.logdet_times[start:idx + 1]
        v = self.logdet_v_vals[start:idx + 1]
        a = self.logdet_a_vals[start:idx + 1]
        tp = self.logdet_t_vals[start:idx + 1]

        self.curve_logdet_v.setData(t, v)
        self.curve_logdet_a.setData(t, a)
        self.curve_logdet_temp.setData(t, tp)

        self.slider_logdet.blockSignals(True)
        self.slider_logdet.setValue(idx)
        self.slider_logdet.blockSignals(False)
        self.lbl_logdet_pos.setText(f"{idx + 1} / {len(self.logdet_times)}")

    def _logdet_step(self):
        if self.logdet_play_index >= len(self.logdet_times):
            self.logdet_timer.stop()
            self.btn_logdet_play_pause.setText("▶ PLAY")
            return
        self._logdet_render_frame(self.logdet_play_index)
        self.logdet_play_index += 1

    def _logdet_toggle_play(self):
        if not self.logdet_times:
            return
        if self.logdet_timer.isActive():
            self.logdet_timer.stop()
            self.btn_logdet_play_pause.setText("▶ PLAY")
        else:
            if self.logdet_play_index >= len(self.logdet_times):
                self.logdet_play_index = 0
            interval_ms = max(20, int(200 / self.logdet_play_speed))
            self.logdet_timer.start(interval_ms)
            self.btn_logdet_play_pause.setText("⏸ PAUSE")

    def _logdet_change_speed(self, combo_idx):
        mapping = {0: 0.5, 1: 1.0, 2: 2.0, 3: 4.0}
        self.logdet_play_speed = mapping.get(combo_idx, 1.0)
        if self.logdet_timer.isActive():
            self.logdet_timer.start(max(20, int(200 / self.logdet_play_speed)))

    def _logdet_seek(self, value):
        self.logdet_play_index = value
        self._logdet_render_frame(value)
    
    def _init_serial_connection(self):
        if serial is not None:
            t = threading.Thread(target=self._read_serial_worker, daemon=True)
            t.start()

    def _resolve_serial_port(self):
        try:
            ports = list(serial.tools.list_ports.comports())
        except Exception:
            ports = []
        if not ports: 
            return SERIAL_PORT
        available = [p.device for p in ports]
        if SERIAL_PORT in available: 
            return SERIAL_PORT
        for p in ports:
            desc = f"{p.description} {p.manufacturer or ''}".upper()
            if any(hint in desc for hint in ESP32_USB_HINTS):
                return p.device
        return ports[0].device if ports else SERIAL_PORT

    def _read_serial_worker(self):
        while True:
            try:
                if self.ser is None or not self.ser.is_open:
                    port_to_use = self._resolve_serial_port()
                    self.ser = serial.Serial(port_to_use, BAUD_RATE, timeout=0.1)
                    self.serial_connected = True

                raw = self.ser.readline()
                if not raw:
                    if time.time() - self.last_packet_time > 2.5:
                        self.packet_loss_flag = True
                    continue

                self.last_packet_time = time.time()
                self.packet_loss_flag = False

                line = raw.decode('utf-8', errors='ignore').strip()
                if not line.startswith("{"): 
                    continue

                try:
                    data = json.loads(line)
                except (json.JSONDecodeError, ValueError):
                    continue

                if data.get("type") == "session_summary":
                    if self.dataset_recording:
                        self._tulis_ringkasan_sesi_dashboard(data)
                    continue

                self.last_raw_line = line
                self.last_raw_data = data
                self.current_v = float(data.get("rms_v", 0.0))
                self.current_a = float(data.get("rms_a", 0.0))
                self.current_temp = float(data.get("temp", 0.0))
                self.current_vx = float(data.get("rms_x", 0.0))
                self.current_vy = float(data.get("rms_y", 0.0))
                self.current_vz = float(data.get("rms_z", 0.0))
                self.current_rpm = float(data.get("rpm", 0.0))
                self.current_d2 = float(data.get("d2", 0.0))
                self.current_status_device = data.get("status", "")

                self.current_health_score = float(data.get("health_score", 100.0))
                self.current_trend = data.get("trend", "Mengumpulkan")
                self.current_servis = data.get("servis_estimasi", "30+ hari")
                self.current_ml_label = data.get("ml_label", "N/A")
                self.current_ml_conf = float(data.get("ml_conf", 0.0))
                self.current_diag_label = data.get("diagnosis", "N/A")
                self.current_diag_conf = float(data.get("diag_conf", 0.0))
                self.current_kurtosis = float(data.get("kurtosis", 3.0))
                self.current_diagnosis_flags = data.get("diag_flags", "Aman")

                self.fft_hz_buffer = data.get("fft_hz", [])
                self.fft_mag_buffer = data.get("fft_mag", [])

                self.tick += 1
                self.time_buffer.append(self.tick)
                self.v_buffer.append(self.current_v)
                self.vx_buffer.append(self.current_vx)
                self.vy_buffer.append(self.current_vy)
                self.vz_buffer.append(self.current_vz)
                self.a_buffer.append(self.current_a)
                self.temp_buffer.append(self.current_temp)
                self.rpm_buffer.append(self.current_rpm)
                self.d2_buffer.append(self.current_d2)

                self.hist_v_buf.append(self.current_v)
                self.hist_a_buf.append(self.current_a)
                self.hist_temp_buf.append(self.current_temp)
                self.hist_vx_buf.append(self.current_vx)
                self.hist_vy_buf.append(self.current_vy)
                self.hist_vz_buf.append(self.current_vz)
                self.hist_rpm_buf.append(self.current_rpm)
                self.hist_d2_buf.append(self.current_d2)

                current_st_lower = (self.current_status_device or "").strip().lower()
                if current_st_lower in ("waspada", "bahaya") and self.last_auto_snapshot_status == "normal":
                    self._trigger_auto_event_snapshot(current_st_lower)
                self.last_auto_snapshot_status = current_st_lower

                if self.recording and self.csv_writer:
                    elapsed = time.perf_counter() - self.record_start_time
                    machine_name = self.machines[self.selected_machine_idx]["name"] if self.selected_machine_idx >= 0 and self.selected_machine_idx < len(self.machines) else "Belum Dipilih"
                    self.csv_writer.writerow([
                        round(elapsed, 3), machine_name,
                        self.current_v, self.current_a, self.current_temp,
                        self.current_vx, self.current_vy, self.current_vz,
                        self.current_rpm, self.current_d2, self.current_status_device,
                        self.current_health_score, self.current_trend, self.current_servis,
                        self.current_ml_label, self.current_diag_label, self.current_kurtosis
                    ])
                    self.csv_file.flush()

                if self.dataset_recording and self.dataset_csv_writer:
                    waktu_dataset = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    baris_dataset = [
                        waktu_dataset,
                        data.get("rms_v"), data.get("rms_x"), data.get("rms_y"), data.get("rms_z"),
                        data.get("rms_a"), data.get("cur"), data.get("cur_raw_adc"),
                        data.get("temp"), data.get("temp_raw"), data.get("rpm"), data.get("snr"),
                        data.get("d2"), data.get("status"),
                        data.get("e_unbalance"), data.get("e_misalign"),
                        data.get("e_bpfo"), data.get("e_bpfi"),
                        data.get("diagnosis"), data.get("diag_conf"),
                        data.get("e_audio_low"), data.get("e_audio_mid"), data.get("e_audio_high"),
                        data.get("audio_diagnosis"), data.get("audio_diag_conf"),
                        data.get("ml_label"), data.get("ml_conf"),
                        data.get("roughness"), data.get("brightness"),
                        data.get("ground_truth"),
                    ]
                    self.dataset_csv_writer.writerow(baris_dataset)
                    self.dataset_csv_file.flush()
            except Exception as e:
                self.serial_connected = False
                self.ser = None
                time.sleep(1)

    def _trigger_auto_event_snapshot(self, status_severity_str):
        try:
            snapshot_filename = os.path.join(LOG_DIR, f"snapshot_{status_severity_str.upper()}_{datetime.now().strftime('%d%m%Y_%H%M%S')}.csv")
            with open(snapshot_filename, 'w', newline='') as sf:
                sw = csv.writer(sf)
                sw.writerow([
                    'snapshot_index', 'rms_v', 'rms_a', 'temp', 'vib_x', 'vib_y', 'vib_z', 
                    'rpm', 'mahalanobis_d2', 'triggered_status'
                ])
                for idx_snap in range(len(self.hist_v_buf)):
                    sw.writerow([
                        idx_snap, self.hist_v_buf[idx_snap], self.hist_a_buf[idx_snap],
                        self.hist_temp_buf[idx_snap], self.hist_vx_buf[idx_snap],
                        self.hist_vy_buf[idx_snap], self.hist_vz_buf[idx_snap],
                        self.hist_rpm_buf[idx_snap], self.hist_d2_buf[idx_snap],
                        status_severity_str
                    ])
        except Exception as ex:
            print(f"Auto-snapshot error: {ex}")

    def _send_command(self, cmd_char):
        if self.ser is not None and self.ser.is_open:
            try:
                self.ser.write(cmd_char.encode())
            except Exception as e:
                print(f"Gagal kirim command {cmd_char}: {e}")
            
    def _apply_header_alarm_state(self, status_key):
        color = {"bahaya": COL_BAD, "waspada": COL_WARN, "diam": COL_IDLE}.get(status_key, COL_ACCENT)
        if self.packet_loss_flag:
            color = COL_WARN
        self.header_frame.setStyleSheet(f"background-color: {COL_HEADER_BG}; border-radius: 3px; border-bottom: 2px solid {color};")

    def _evaluate_diagnosis(self, v, temp, device_status=""):
        status_key = (device_status or "").strip().lower()
        self._apply_header_alarm_state(status_key)
        
        status_map = {
            "diam": ("● MOTOR DIAM", COL_IDLE),
            "bahaya": ("● DEVIASI BAHAYA", COL_BAD),
            "waspada": ("● STATUS WASPADA", COL_WARN),
            "normal": ("● SYSTEM ONLINE", COL_OK),
            "warming": ("● WARMING UP", "#888888"),
            "notcalibrated": ("● TANPA KALIBRASI / DIRECT", COL_OK),
            "sensorfault": ("● SENSOR FAULT", COL_WARN),
        }
        
        txt, col = status_map.get(status_key, ("● SYSTEM ONLINE", COL_ACCENT))
        if self.packet_loss_flag:
            txt, col = "● PACKET LOSS WARNING", COL_WARN
        self.lbl_sys_status.setText(txt)
        self.lbl_sys_status.setStyleSheet(f"font-size: 8px; font-weight: bold; color: {col};")

    def _show_debug_info(self):
        port_info = self.ser.port if (self.ser is not None) else "(belum ada koneksi)"
        selected_machine_name = self.machines[self.selected_machine_idx]["name"] if self.selected_machine_idx >= 0 and self.selected_machine_idx < len(self.machines) else "Belum dipilih"
        active_slot = self.machines[self.selected_machine_idx].get("slot_char", "-") if self.selected_machine_idx >= 0 and self.selected_machine_idx < len(self.machines) else "N/A"
        is_no_calib = self.machines[self.selected_machine_idx].get("no_calib", False) if self.selected_machine_idx >= 0 and self.selected_machine_idx < len(self.machines) else False
        info = (
            f"Status koneksi   : {'TERSAMBUNG' if self.serial_connected else 'TIDAK TERSAMBUNG'}\n"
            f"Port serial      : {port_info}\n"
            f"Baud rate        : {BAUD_RATE}\n"
            f"Packet Loss Flag : {'TERDETEKSI' if self.packet_loss_flag else 'Aman'}\n"
            f"Mesin/Slot Aktif : {selected_machine_name} (Slot: {active_slot})\n"
            f"Mode Kalibrasi   : {'TANPA KALIBRASI (Direct)' if is_no_calib else 'Dengan Kalibrasi Baseline'}\n"
            f"Baris JSON akhir : {self.last_raw_line or '(belum ada data)'}\n"
            f"Vib/Snd/Tmp      : {self.current_v}, {self.current_a}, {self.current_temp}\n"
            f"RPM / D²         : {self.current_rpm}, {self.current_d2}\n"
            f"Status firmware  : {self.current_status_device or '-'}"
        )
        QMessageBox.information(self, "DEBUG - Info Koneksi & Protokol", info)

    def _render_session_summary(self):
        self.lbl_session_summary.setText(
            f"Sesi: {self.session_sample_count} sample | "
            f"RPM rata-rata: {(self.session_rpm_sum / self.session_sample_count) if self.session_sample_count else 0.0:.1f} | "
            f"D² max: {self.session_d2_max:.2f} | Kondisi terparah: {self.session_worst_status}"
        )

    def _update_gui(self):
        now_dt = datetime.now()
        self.time_lbl.setText(now_dt.strftime("%H:%M:%S"))
        self.date_lbl.setText(now_dt.strftime("%d/%m/%Y"))

        self.lbl_conn_dot.setStyleSheet(
            f"font-size: 10px; font-weight: bold; color: {COL_OK if (self.serial_connected and not self.packet_loss_flag) else COL_WARN};"
        )

        if self.current_v is not None:
            self.tick += 1


            self.curve_v.setData(list(self.time_buffer), list(self.v_buffer))
            self.curve_a.setData(list(self.time_buffer), list(self.a_buffer))
            self.curve_temp.setData(list(self.time_buffer), list(self.temp_buffer))

            self.curve_rpm.setData(list(self.time_buffer), list(self.rpm_buffer))
            self.curve_d2.setData(list(self.time_buffer), list(self.d2_buffer))

            if len(self.fft_hz_buffer) > 0 and len(self.fft_hz_buffer) == len(self.fft_mag_buffer):
                self.curve_fft.setData(self.fft_hz_buffer, self.fft_mag_buffer)

            self.lbl_hs_val.setText(f"{self.current_health_score:.1f}%")
            self.prog_hs.setValue(int(self.current_health_score))
            
            self.lbl_ml_val.setText(f"{self.current_ml_label} ({self.current_ml_conf*100:.0f}%)")
            self.lbl_diag_val.setText(f"{self.current_diag_label} ({self.current_diag_conf*100:.0f}%)")
            
            self.lbl_kurt_val.setText(f"{self.current_kurtosis:.2f}")
            self.lbl_flags_val.setText(f"{self.current_diagnosis_flags}")

            self.lbl_val_v.setText(f"{self.current_v:.2f} G")

            if self.current_vx is not None:
                self.lbl_val_vxyz.setText(
                    f"(X: {self.current_vx:.2f} | Y: {self.current_vy:.2f} | Z: {self.current_vz:.2f} G)"
                )
            self.lbl_val_a.setText(f"{self.current_a:.1f} dB")
            self.lbl_val_temp.setText(f"{self.current_temp:.1f} °C")
            if self.current_rpm is not None:
                self.lbl_val_rpm.setText(f"{self.current_rpm:.0f}")
            if self.current_d2 is not None:
                self.lbl_val_d2.setText(f"{self.current_d2:.2f}")

            rpm_txt = f"{self.current_rpm:.0f}" if self.current_rpm is not None else "--"
            d2_txt = f"{self.current_d2:.2f}" if self.current_d2 is not None else "--"
            self.lbl_proc_snapshot.setText(
                f"Live | Vib: {self.current_v:.2f} G | Snd: {self.current_a:.1f} dB | "
                f"Tmp: {self.current_temp:.1f} °C | RPM: {rpm_txt} | D²: {d2_txt}"
            )

            self.gauge_v.set_value(self.current_v, status_key)
            self.gauge_s.set_value(self.current_a, status_key)
            self.gauge_t.set_value(self.current_temp, status_key)

            machine_name = self.machines[self.selected_machine_idx]["name"] if self.selected_machine_idx >= 0 and self.selected_machine_idx < len(self.machines) else "Belum dipilih"
            self.lbl_sum_machine.setText(f"Target: {machine_name}")

            self._evaluate_diagnosis(self.current_v, self.current_temp, self.current_status_device)

            if self.tick != self.last_processed_tick:
                self.last_processed_tick = self.tick
                self.session_sample_count += 1
                if self.current_rpm is not None:
                    self.session_rpm_sum += self.current_rpm
                if self.current_d2 is not None:
                    self.session_d2_max = max(self.session_d2_max, self.current_d2)

                if status_key == "waspada":
                    self.session_waspada_count += 1
                elif status_key == "bahaya":
                    self.session_bahaya_count += 1

                if status_key in STATUS_SEVERITY:
                    if STATUS_SEVERITY[status_key] > STATUS_SEVERITY.get(self.session_worst_status.lower(), 0):
                        self.session_worst_status = self.current_status_device.capitalize()

                if status_key in ("waspada", "bahaya"):
                    ts = datetime.now().strftime("%H:%M:%S")
                    event_txt = (
                        f"[{ts}] {self.current_status_device.upper()} — RPM {rpm_txt}, D² {d2_txt}, "
                        f"Vib {self.current_v:.2f}G, Tmp {self.current_temp:.1f}°C"
                    )
                    self.anomaly_events.append(event_txt)
                    if self.list_anomali.count() == 1 and self.list_anomali.item(0).text().startswith("Tidak ada"):
                        self.list_anomali.clear()
                    self.list_anomali.addItem(event_txt)
                    self.list_anomali.scrollToBottom()

                self._render_session_summary()
        elif not self.serial_connected:
            self.lbl_sys_status.setText("● MENCARI PERANGKAT (COM3)...")
            self.lbl_sys_status.setStyleSheet(f"font-size: 8px; font-weight: bold; color: {COL_WARN};")
        else:
            self.lbl_sys_status.setText("● TERSAMBUNG — MENUNGGU DATA JSON...")
            self.lbl_sys_status.setStyleSheet(f"font-size: 8px; font-weight: bold; color: {COL_ACCENT};")
            self.lbl_proc_snapshot.setText("Menunggu data serial dari ESP32...")

    def _toggle_dataset_recording(self):
        if self.selected_machine_idx < 0:
            QMessageBox.warning(self, "Perhatian", "Silakan pilih Target Mesin terlebih dahulu!")
            return

        if not self.dataset_recording:
            m = self.machines[self.selected_machine_idx]
            try:
                slot_int = int(m["slot_char"])
            except (ValueError, TypeError):
                slot_int = -1
            slot_label = SLOT_LABEL_UNTUK_NAMA_FILE.get(slot_int, f"slot{m['slot_char']}")

            kondisi_idx = self.cmb_rec_kondisi.currentIndex()
            huruf_kondisi, label_kondisi = KONDISI_REKAM_OPSI[kondisi_idx]

            self._send_command(huruf_kondisi)

            ts = datetime.now().strftime("%Y%m%d_%H%M")
            base_name = f"vibris_{ts}_{slot_label}_{label_kondisi}"
            csv_path = os.path.join(DATASET_DIR, f"{base_name}.csv")
            self.dataset_summary_path = os.path.join(DATASET_DIR, f"{base_name}_ringkasan_sesi.txt")

            try:
                file_baru = not os.path.exists(csv_path)
                self.dataset_csv_file = open(csv_path, 'a', newline='', encoding='utf-8')
                self.dataset_csv_writer = csv.writer(self.dataset_csv_file)
                if file_baru:
                    self.dataset_csv_writer.writerow(KOLOM_DATASET)
                self.dataset_recording = True
                self.btn_toggle_dataset_rec.setText("⏹ BERHENTI SIMPAN DATASET")
                self.btn_toggle_dataset_rec.setStyleSheet(f"background-color: {COL_BAD}; color: #ffffff; font-weight: bold; font-size: 9px; height: 24px; border-radius: 3px;")
                self.lbl_dataset_rec_status.setText(f"● MENULIS -> {os.path.basename(csv_path)}")
            except Exception as e:
                QMessageBox.warning(self, "Gagal", f"Gagal membuat file Dataset: {e}")
        else:
            self.dataset_recording = False
            if self.dataset_csv_file:
                self.dataset_csv_file.close()
                self.dataset_csv_file = None
                self.dataset_csv_writer = None
            self.btn_toggle_dataset_rec.setText("💾 SIMPAN KE DATASET (Format Logger)")
            self.btn_toggle_dataset_rec.setStyleSheet("background-color: #2f7d4f; color: #ffffff; font-weight: bold; font-size: 9px; height: 24px; border-radius: 3px;")
            self.lbl_dataset_rec_status.setText("● Berkas Dataset Tersimpan")

    def _tulis_ringkasan_sesi_dashboard(self, data):
        if not self.dataset_summary_path:
            return
        waktu = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        teks = (
            f"\n{'=' * 60}\n"
            f"[RINGKASAN SESI CEK] {waktu}\n"
            f"  Slot mesin          : #{data.get('slot')}\n"
            f"  Kesimpulan dominan  : {data.get('dominant')}\n"
            f"  Durasi sesi         : {data.get('duration_ms')} ms "
            f"({data.get('total_samples')} sample)\n"
            f"  Normal={data.get('n_normal')}  Waspada={data.get('n_waspada')}  "
            f"Bahaya={data.get('n_bahaya')}  Diam={data.get('n_diam')}\n"
            f"  Diagnosis terdeteksi: Unbalance={data.get('n_unbalance')} "
            f"Misalign={data.get('n_misalign')} BPFO={data.get('n_bpfo')} "
            f"BPFI={data.get('n_bpfi')}\n"
            f"  Rata-rata Health Score : {data.get('avg_health')}\n"
            f"  Suhu awal -> akhir sesi: {data.get('temp_start')}C -> "
            f"{data.get('temp_end')}C (delta {data.get('temp_delta')}C)\n"
            f"{'=' * 60}\n"
        )
        try:
            with open(self.dataset_summary_path, "a", encoding="utf-8") as f:
                f.write(teks)
        except Exception as e:
            print(f"Gagal menulis ringkasan sesi Dataset: {e}")

    def _toggle_recording(self):
        if self.selected_machine_idx < 0:
            QMessageBox.warning(self, "Perhatian", "Silakan pilih Target Mesin terlebih dahulu!")
            return
            
        if not self.recording:
            filename = os.path.join(LOG_DIR, f"log_{datetime.now().strftime('%d%m%Y_%H%M%S')}.csv")
            try:
                self.csv_file = open(filename, 'w', newline='')
                self.csv_writer = csv.writer(self.csv_file)
                self.csv_writer.writerow([
                    'timestamp', 'machine_type', 'rms_v', 'rms_a', 'temp', 'vib_x', 'vib_y', 'vib_z', 
                    'rpm', 'mahalanobis_d2', 'status', 'health_score', 'trend', 'servis_est', 'ml_label', 'diag_label', 'kurtosis'
                ])
                self.record_start_time = time.perf_counter()
                self.recording = True
                self.btn_toggle_rec.setText("BERHENTI RECORDING")
                self.btn_toggle_rec.setStyleSheet(f"background-color: {COL_BAD}; color: #ffffff; font-weight: bold; font-size: 9px; height: 24px;")
                self.lbl_rec_status.setText(f"● MENULIS -> {os.path.basename(filename)}")
            except Exception as e:
                print(f"Gagal membuat file log: {e}")
        else:
            self.recording = False
            if self.csv_file:
                self.csv_file.close()
                self.csv_file = None
            self.btn_toggle_rec.setText("MULAI RECORDING")
            self.btn_toggle_rec.setStyleSheet("background-color: #cfcfcf; color: #000000; font-weight: bold; font-size: 9px; height: 24px;")
            self.lbl_rec_status.setText("● Berkas Tersimpan")
            self._refresh_log_list()

    def _refresh_log_list(self):
        self.log_list.clear()
        if os.path.isdir(LOG_DIR):
            for file in sorted(os.listdir(LOG_DIR), reverse=True):
                if file.endswith(".csv"):
                    self.log_list.addItem(file)

    def _open_log_detail(self, item):
        if item is None:
            return
        filename = item.text()
        filepath = os.path.join(LOG_DIR, filename)

        self.logdet_timer.stop()
        self.logdet_play_index = 0
        self.btn_logdet_play_pause.setText("▶ PLAY")
        self.lbl_logdet_title.setText(f"HASIL DETEKSI REKAMAN: {filename}")

        (self.logdet_times, self.logdet_v_vals, self.logdet_a_vals,
         self.logdet_t_vals) = self._logdet_load_csv(filepath)

        self.slider_logdet.setMaximum(max(0, len(self.logdet_times) - 1))
        self.slider_logdet.setValue(0)
        if self.logdet_times:
            self._logdet_render_frame(0)
        else:
            self.curve_logdet_v.clear()
            self.curve_logdet_a.clear()
            self.curve_logdet_temp.clear()
            self.lbl_logdet_pos.setText("0 / 0")
        self._logdet_render_diagnosis_summary()

        self._change_page(3)

    def _open_log_detail_from_button(self):
        current_item = self.log_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "Perhatian", "Pilih file rekaman (.csv) terlebih dahulu!")
            return
        self._open_log_detail(current_item)

    def _delete_selected_log(self):
        current_item = self.log_list.currentItem()
        if not current_item:
            return
        filename = current_item.text()
        filepath = os.path.join(LOG_DIR, filename)
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
            self._refresh_log_list()
            self.lbl_rec_status.setText("● BERKAS DIHAPUS")
        except Exception as e:
            print(f"Gagal menghapus file: {e}")

    def _export_selected_log_to_excel(self):
        current_item = self.log_list.currentItem()
        if not current_item:
            QMessageBox.information(self, "Pilih Rekaman Dulu", "Klik salah satu berkas rekaman di daftar dulu.")
            return

        if openpyxl is None:
            QMessageBox.critical(self, "Library Belum Terinstall", "Fitur export ke Excel butuh library 'openpyxl'.")
            return

        filename = current_item.text()
        csv_path = os.path.join(LOG_DIR, filename)
        xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"

        try:
            with open(csv_path, 'r') as f:
                rows = list(csv.reader(f))

            if not rows:
                QMessageBox.warning(self, "Berkas Kosong", "Berkas rekaman ini tidak memiliki data.")
                return

            header, data_rows = rows[0], rows[1:]

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Sensor ESP32"

            header_fill = PatternFill(start_color="1c1e22", end_color="1c1e22", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            ws.append(header)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            status_col_idx = None
            for i, col_name in enumerate(header):
                if col_name.strip().lower() == "status":
                    status_col_idx = i
                    break

            status_fill = {
                "diam":    PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid"),
                "normal":  PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid"), 
                "waspada": PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid"),  
                "bahaya":  PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid"),   
            }

            for row in data_rows:
                converted_row = []
                for value in row:
                    try:
                        converted_row.append(float(value))
                    except (ValueError, TypeError):
                        converted_row.append(value)
                ws.append(converted_row)

                if status_col_idx is not None:
                    status_val = str(row[status_col_idx]).strip().lower()
                    fill = status_fill.get(status_val)
                    if fill:
                        ws.cell(row=ws.max_row, column=status_col_idx + 1).fill = fill

            for col_cells in ws.columns:
                max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
                ws.column_dimensions[col_cells[0].column_letter].width = max_len + 2

            ws.freeze_panes = "A2"  
            wb.save(xlsx_path)

            QMessageBox.information(
                self, "Export Berhasil",
                f"Data berhasil diexport ke:\n{xlsx_path}\n\nTotal {len(data_rows)} baris data."
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Gagal", f"Terjadi kesaBlahan saat export:\n{e}")

class GlobalEscHandler(QApplication):
    def notify(self, receiver, event):
        if event.type() == QEvent.KeyPress and event.key() == Qt.Key_Escape:
            for widget in self.topLevelWidgets():
                if isinstance(widget, Dashboard):
                    if widget.csv_file:
                        widget.csv_file.close()
                    widget.close()
                    return True
        return super().notify(receiver, event)

if __name__ == '__main__':
    app = GlobalEscHandler(sys.argv)
    db = Dashboard()
    db.show()
    db.raise_()
    db.activateWindow()
    sys.exit(app.exec_())
